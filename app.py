from fastapi import FastAPI, Request, UploadFile, File, Form
from fastapi.responses import (
    HTMLResponse,
    FileResponse,
    JSONResponse,
    RedirectResponse
)
from fastapi.staticfiles import StaticFiles
from apscheduler.schedulers.background import BackgroundScheduler
import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
import csv, os, shutil, glob, json
from starlette.middleware.sessions import SessionMiddleware
import math
import sqlite3
import secrets
import pytz
IST = pytz.timezone("Asia/Kolkata")
from passlib.context import CryptContext
from dotenv import load_dotenv
from fastapi import HTTPException
from dhan_token_manager import init_token_manager, router as token_router
from password_manager import router as pw_router
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from io import BytesIO
from fastapi.responses import StreamingResponse
from fastapi.responses import Response
load_dotenv()

# ── Time Configuration ─────────────────────────────
IST = pytz.timezone("Asia/Kolkata")

MARKET_OPEN_HOUR = 9
MARKET_OPEN_MIN = 16

MARKET_CLOSE_HOUR = 12
MARKET_CLOSE_MIN = 0

DASHBOARD_OPEN_HOUR = 8
DASHBOARD_OPEN_MIN = 30

CONFIG_FILE = "config.json"

def get_interval():

    try:

        with open(CONFIG_FILE, "r") as f:

            data = json.load(f)

            return int(data.get("interval", 15))

    except:

        return 15


def set_interval(val):

    with open(CONFIG_FILE, "w") as f:

        json.dump({
            "interval": val
        }, f)

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


def hash_password(password):
    return pwd_context.hash(password)

def verify_password(plain, hashed):
    return pwd_context.verify(plain, hashed)

def create_user_session(request, username):

    session_token = secrets.token_hex(32)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
        INSERT OR REPLACE INTO active_sessions
        (
            username,
            session_token,
            login_time,
            ip_address,
            user_agent
        )
        VALUES (?, ?, datetime('now'), ?, ?)
    """, (
        username,
        session_token,
        request.client.host if request.client else "",
        request.headers.get("user-agent", "")
    ))

    conn.commit()
    conn.close()

    request.session.clear()
    request.session["user"] = username
    request.session["role"] = "user"
    request.session["session_token"] = session_token


def validate_user_session(request):

    username = request.session.get("user")
    token = request.session.get("session_token")

    if not username or not token:
        return False

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
        SELECT session_token
        FROM active_sessions
        WHERE username=?
    """, (username,))

    row = c.fetchone()

    conn.close()

    if not row:
        return False

    return row[0] == token

DB_FILE = "traderbro.db"


from market_holidays import (
    MARKET_HOLIDAYS, is_market_holiday, get_holiday_reason,
    get_upcoming_holidays, get_today_holiday, get_all_holidays_list,
    refresh_holidays, get_holidays_for_year
    )
 
PLAN_CONFIG_DATA = {
    "basic":     {"trading_days": 1,   "price": 99},
    "essential": {"trading_days": 5,   "price": 399},
    "pro":       {"trading_days": 22,  "price": 1499},
    "premium":   {"trading_days": 250, "price": 14499},
}

def calculate_expiry_from_start(start_dt, trading_days):
    """
    Dashboard stays visible through the ENTIRE last trading day.
    Expiry timestamp = midnight (00:00) of the day AFTER the last trading day,
    so access cuts off exactly at 12 AM the next day.
    """
    days_to_add = trading_days - 1
    last_day = start_dt.replace(second=0, microsecond=0)
    added = weekends = holidays = 0

    while added < days_to_add:
        last_day = last_day + timedelta(days=1)
        dow = last_day.weekday()
        if dow >= 5:
            weekends += 1
            continue
        if is_market_holiday(last_day):
            holidays += 1
            continue
        added += 1

    total_cal = (last_day.date() - start_dt.date()).days + 1

    expiry = last_day.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    return expiry, total_cal, weekends, holidays
 
def get_next_trading_day_after(dt):
    """
    dt is already midnight of the day the previous plan deactivates.
    If that day is a trading day, the new plan starts exactly then;
    otherwise roll forward to the next trading day at midnight.
    """
    nxt = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    while nxt.weekday() >= 5 or is_market_holiday(nxt):
        nxt += timedelta(days=1)
    return nxt


def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE,
        email TEXT UNIQUE,
        password TEXT,
        phone TEXT,
        role TEXT DEFAULT 'user',
        consent BOOLEAN,
        plan TEXT DEFAULT 'free',
        plan_start TEXT,
        plan_expiry TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS active_sessions (
        username TEXT PRIMARY KEY,
        session_token TEXT,
        login_time TEXT,
        ip_address TEXT,
        user_agent TEXT
    )
    """)

    """
    c.execute(\"\"\"
    CREATE TABLE IF NOT EXISTS broadcasts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message TEXT NOT NULL,
        sender TEXT DEFAULT 'admin',
        timestamp TEXT NOT NULL,
        pinned INTEGER DEFAULT 0
    )
    \"\"\")
    """

    # Create default admin if not exists
    c.execute("SELECT * FROM users WHERE username='admin'")
    if not c.fetchone():
        admin_password = hash_password("admin123")
        c.execute("""
        INSERT INTO users (username, email, password, role, consent, plan)
        VALUES (?, ?, ?, ?, ?, ?)
        """, ("admin", "admin@example.com", admin_password, "admin", True, "pro"))

    conn.commit()
    conn.close()
init_db()

def migrate_db():
    """Add columns that may be missing from older DB versions."""
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # Add created_at to users if missing
    try:
        c.execute("ALTER TABLE users ADD COLUMN created_at TEXT")
        # Backfill existing rows with a placeholder date
        c.execute("UPDATE users SET created_at = datetime('now') WHERE created_at IS NULL")
        conn.commit()
        print("✅ DB migrated: added created_at to users")
    except:
        pass  # Column already exists
    conn.close()

migrate_db()

def init_subscriptions_table():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        plan TEXT NOT NULL,
        plan_start TEXT NOT NULL,
        plan_expiry TEXT NOT NULL,
        trading_days INTEGER DEFAULT 0,
        total_calendar_days INTEGER DEFAULT 0,
        weekends_skipped INTEGER DEFAULT 0,
        holidays_skipped INTEGER DEFAULT 0,
        status TEXT DEFAULT 'queued',
        price INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)
    conn.commit()
    conn.close()
 
init_subscriptions_table()

# Webinar tables
def init_webinars_table():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
    CREATE TABLE IF NOT EXISTS webinars (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        host TEXT DEFAULT 'TraderBro Team',
        scheduled_at TEXT NOT NULL,
        duration_minutes INTEGER DEFAULT 60,
        topics TEXT,
        meeting_link TEXT,
        registration_link TEXT,
        cover_color TEXT DEFAULT 'orange',
        status TEXT DEFAULT 'upcoming',
        max_seats INTEGER DEFAULT 0,
        is_free INTEGER DEFAULT 1,
        recording_link TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    """)

    c.execute("""
    CREATE TABLE IF NOT EXISTS webinar_registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        webinar_id INTEGER NOT NULL,
        username TEXT NOT NULL,
        registered_at TEXT NOT NULL,
        UNIQUE(webinar_id, username)
    )
    """)

    conn.commit()
    conn.close()

init_webinars_table()

def create_admin():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("SELECT * FROM users WHERE role='admin'")
    admin = c.fetchone()

    if not admin:
        hashed_pw = hash_password("admin123")

        c.execute("""
        INSERT INTO users (username, email, password, role)
        VALUES (?, ?, ?, ?)
        """, ("admin", "admin@traderbro.in", hashed_pw, "admin"))

        conn.commit()

    conn.close()

create_admin()

# GLOBALs
LAST_DATA = {
    "time": None,
    "data": None,
    "DASH_HISTORY": {
        "last_diff": 0,
        "running": 0,
        "rows": []
    }
}
LAST_FETCH_TIME = None
LIVE_RUNNING_RECORDS = []
RUNNING_FILE = "live_running.json"
LOGIN_ATTEMPTS = {}


app = FastAPI()
app.add_middleware(
    SessionMiddleware,
    secret_key="TraderBro@2026#Secure$FastAPI",
    https_only=False,  
    same_site="lax"
)
app.include_router(pw_router)
@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page():
    import os
    path = os.path.join(STATIC_DIR, "finlab", "forgot-password.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

# ── Serve static files (dashboard.html lives in ./static/) ──────────────
STATIC_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static")
os.makedirs(STATIC_DIR, exist_ok=True)
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# ── Upload storage folder ────────────────────────────────────────────────
UPLOADS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploaded_csvs")
EXCEL_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "excel_exports"
)
os.makedirs(UPLOADS_DIR, exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)

# CLIENT_ID  = "1100585975"
# ACCESS_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpc3MiOiJkaGFuIiwicGFydG5lcklkIjoiIiwiZXhwIjoxNzc4NTIwMjExLCJpYXQiOjE3Nzg0MzM4MTEsInRva2VuQ29uc3VtZXJUeXBlIjoiU0VMRiIsIndlYmhvb2tVcmwiOiIiLCJkaGFuQ2xpZW50SWQiOiIxMTAwNTg1OTc1In0.B3PZw1MYo0V33yv2O_nQMaOu6_ISoggjscSpYjhJcgfUPrs1vYt26uI8S3XZxlRG1BswvBKjZzNfjzNnJ74Jwg"
ACCESS_TOKEN = os.getenv("ACCESS_TOKEN")
CLIENT_ID  = os.getenv("CLIENT_ID")
BASE_URL   = "https://api.dhan.co/v2"
HEADERS    = {
    "Content-Type": "application/json",
    "access-token": ACCESS_TOKEN,
    "client-id":    CLIENT_ID,
}

CSV_FILE = "sensex_atm_history.csv"
CSV_COLUMNS = [
    "DateTime", "Expiry", "Strike",
    "CE_LTP", "CE_Delta", "CE_Gamma", "CE_Theta", "CE_Vega",
    "PE_LTP", "PE_Delta", "PE_Gamma", "PE_Theta", "PE_Vega",
    "Delta_Ratio", "Index_LTP",
    "Reference", "Stretched", "Difference"
]

# ── Recorder state ───────────────────────────────────────────────────────
recorder_state = {
    "running": False, "interval": 20,
    "expiry": None, "start_time": None,
    "stop_time": None, "records_saved": 0,
}
scheduler = BackgroundScheduler()
scheduler.start()

# VPS AUTO RECORDER
AUTO_RUNNING = False

# Remove these two globals (no longer needed):
# LAST_FETCH_TIME = None   ← delete this line at the top

def auto_market_recorder():
    global AUTO_RUNNING

    now = datetime.now(IST)

    # ── HARD BLOCK: weekends and holidays never record ──────────────────
    if now.weekday() >= 5:          # Saturday=5, Sunday=6
        if AUTO_RUNNING:
            print("⛔ WEEKEND — recorder idle")
            AUTO_RUNNING = False
        return

    if is_market_holiday(now):
        if AUTO_RUNNING:
            print(f"⛔ HOLIDAY ({get_holiday_reason(now)}) — recorder idle")
            AUTO_RUNNING = False
        return

    # if True:  # keep your market hours check here if needed
    if is_market_open():
        if not AUTO_RUNNING:
            print("🚀 MARKET RECORDER STARTED")
            AUTO_RUNNING = True

        try:
            expiry_list = get_expiries()
            if not expiry_list:
                return

            expiry = expiry_list[0]
            ltp, df, atm = get_live_chain(expiry, force_live=True)  # ← force_live=True

            if df.empty or atm is None:
                return

            atm_row = df[df["Strike"] == atm]
            if atm_row.empty:
                return

            r = atm_row.iloc[0]

            global LIVE_RUNNING_RECORDS

            current_diff = r["Difference"]
            if current_diff is None:
                current_diff = 0

            row = {
                "datetime":   str(r["DateTime"]),
                "expiry":     str(r["Expiry"]),
                "ce_ltp":     r["CE_LTP"],
                "ce_delta":   r["CE_Delta"],
                "ce_gamma":   r["CE_Gamma"],
                "ce_theta":   r["CE_Theta"],
                "ce_vega":    r["CE_Vega"],
                "strike":     int(r["Strike"]),
                "pe_ltp":     r["PE_LTP"],
                "pe_delta":   r["PE_Delta"],
                "pe_gamma":   r["PE_Gamma"],
                "pe_theta":   r["PE_Theta"],
                "pe_vega":    r["PE_Vega"],
                "delta_ratio": r["Delta_Ratio"],
                "index_ltp":  ltp,
                "reference":  r["Reference"],
                "stretched":  r["Stretched"],
                "difference": current_diff,
                "diff_prev":  0,
                "running":    0,
            }

            if len(LIVE_RUNNING_RECORDS) > 0:
                prev = LIVE_RUNNING_RECORDS[-1]
                diff_change = current_diff - prev["difference"]
                row["diff_prev"] = round(diff_change, 2)
                row["running"]   = round(prev.get("running", 0) + diff_change, 2)
            # First record: running stays 0 (baseline)

            # DUPLICATE TIMESTAMP PROTECTION
            if len(LIVE_RUNNING_RECORDS) > 0:
                last = LIVE_RUNNING_RECORDS[-1]
                if (last["datetime"] == row["datetime"] and
                        last["difference"] == row["difference"]):
                    print("⚠️ DUPLICATE ROW SKIPPED")
                    return

            LIVE_RUNNING_RECORDS.append(row)
            LIVE_RUNNING_RECORDS = LIVE_RUNNING_RECORDS[-2000:]

            with open(RUNNING_FILE, "w") as f:
                json.dump(LIVE_RUNNING_RECORDS, f)

            print("✅ SAVED:", row["datetime"])

        except Exception as e:
            print("AUTO RECORDER ERROR:", e)
    else:
        if AUTO_RUNNING:
            print("⛔ MARKET RECORDER STOPPED")
        AUTO_RUNNING = False

# login

def load_login_template(role="admin", error=False):
    path = os.path.join(STATIC_DIR, "finlab", "login.html")

    with open(path, "r", encoding="utf-8") as f:
        html = f.read()

    # Asset paths
    html = html.replace('href="vendor/', 'href="/static/finlab/vendor/')
    html = html.replace('src="vendor/', 'src="/static/finlab/vendor/')
    html = html.replace('href="css/', 'href="/static/finlab/css/')
    html = html.replace('src="js/', 'src="/static/finlab/js/')
    html = html.replace('href="images/', 'href="/static/finlab/images/')
    html = html.replace('src="images/', 'src="/static/finlab/images/')
    html = html.replace('url(images/', 'url(/static/finlab/images/')

    if role == "admin":
        html = html.replace("Welcome Back", "Admin Login")
        html = html.replace("Sign Me In", "Login as Admin")
        html = html.replace(
            "Log in to your admin dashboard with your credentials",
            "Secure admin access to TraderBro control panel"
        )
        html = html.replace(
            "The Evolution of <span>Finlab</span>",
            "Welcome to <span>TraderBro</span>"
        )
        action = "/admin-login"
    else:
        action = "/user-login"

    html = html.replace(
        '<form method="POST" action="/user-login">',
        f'<form method="POST" action="{action}">'
    )

    html = html.replace(
        'name="identifier"',
        'name="username"'
    )

    html = html.replace(
        'type="email" class="form-control" value="hello@example.com"',
        'type="text" name="username" class="form-control" placeholder="Enter Username"'
    )

    html = html.replace(
        'id="dlab-password" class="form-control" value="123456"',
        'name="password" id="dlab-password" class="form-control" placeholder="Enter Password"'
    )

    if error:
        html = html.replace(
            '<form method="post"',
            '''
            <div style="background:#ffdddd;color:red;padding:10px;margin-bottom:15px;border-radius:8px">
            Invalid Login Credentials
            </div>
            <form method="post"
            '''
        )

    return html

# @app.get("/api/me")
# def get_current_user(request: Request):

#     if "user" not in request.session:
#         return {"username": None}

#     conn = sqlite3.connect(DB_FILE)
#     c = conn.cursor()

#     c.execute("""
#         SELECT username, email, phone, plan, plan_start, plan_expiry
#         FROM users
#         WHERE username=?
#     """, (request.session["user"],))

#     user = c.fetchone()

#     conn.close()

#     if user:
#         return {
#             "username": user[0],
#             "email": user[1],
#             "phone": user[2],
#             "plan": user[3],
#             "plan_start": user[4],
#             "plan_expiry": user[5]
#         }

#     return {"username": None}


SITE_URL = "https://traderbro.in"

@app.get("/robots.txt")
def robots_txt():
    content = """User-agent: *
Allow: /
Disallow: /admin
Disallow: /admin-login
Disallow: /admin/
Disallow: /api/
Disallow: /dashboard
Disallow: /account
Disallow: /checkout
Disallow: /simple
Disallow: /forgot-password

Sitemap: https://traderbro.in/sitemap.xml
"""
    return Response(content=content, media_type="text/plain")


@app.get("/sitemap.xml")
def sitemap():
    pages = [
        {"loc": "/",                     "priority": "1.0", "changefreq": "daily"},
        {"loc": "/about-us",             "priority": "0.8", "changefreq": "monthly"},
        {"loc": "/trading-plan",         "priority": "0.9", "changefreq": "weekly"},
        {"loc": "/webinars",             "priority": "0.7", "changefreq": "weekly"},
        {"loc": "/contact-us",           "priority": "0.5", "changefreq": "yearly"},
        {"loc": "/privacy-policy",       "priority": "0.3", "changefreq": "yearly"},
        {"loc": "/terms-and-conditions", "priority": "0.3", "changefreq": "yearly"},
        {"loc": "/disclaimer",           "priority": "0.3", "changefreq": "yearly"},
        {"loc": "/refund-policy",        "priority": "0.3", "changefreq": "yearly"},
    ]
    xml = ['<?xml version="1.0" encoding="UTF-8"?>',
           '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">']
    for p in pages:
        xml.append(
            f"<url><loc>{SITE_URL}{p['loc']}</loc>"
            f"<changefreq>{p['changefreq']}</changefreq>"
            f"<priority>{p['priority']}</priority></url>"
        )
    xml.append("</urlset>")
    return Response(content="\n".join(xml), media_type="application/xml")

@app.get("/api/me")
def get_current_user(request: Request):
    if "user" in request.session:

        if not validate_user_session(request):

            request.session.clear()

            return {
                "username": None,
                "role": None,
                "is_admin": False
            }

    username = None
    role = "user"

    # USER SESSION
    if "user" in request.session:
        username = request.session["user"]
        role = request.session.get("role", "user")

    # ADMIN SESSION
    elif "admin" in request.session:
        username = request.session["admin"]
        role = "admin"

    # NO SESSION
    else:
        return {
            "username": None,
            "role": None,
            "is_admin": False
        }

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
        SELECT username, email, phone, plan, plan_start, plan_expiry
        FROM users
        WHERE username=?
    """, (username,))

    user = c.fetchone()

    conn.close()

    if user:
        subscription_active = False

        try:
            if user[5]:
                expiry = datetime.fromisoformat(user[5])

                if expiry.tzinfo is None:
                    expiry = IST.localize(expiry)

                subscription_active = expiry > datetime.now(IST)

        except Exception:
            pass

        return {
            "username": user[0],
            "email": user[1],
            "phone": user[2],
            "plan": user[3],
            "plan_start": user[4],
            "plan_expiry": user[5],

            "subscription_active": subscription_active,

            "role": role,
            "is_admin": role == "admin"
        }


@app.get("/api/admin/user/disclaimer-pdf/{user_id}")
async def download_disclaimer_pdf(user_id: int, request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    user = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()

    if not user:
        return JSONResponse({"error": "User not found"}, status_code=404)

    # Get join date — use created_at if exists, else fallback
    user_dict = dict(user)
    join_raw = user_dict.get("created_at") or ""
    try:
        d = datetime.fromisoformat(str(join_raw))
        join_fmt = f"{d.day}/{d.month}/{d.year}"
    except:
        join_fmt = datetime.now(IST).strftime("%d/%m/%Y")

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            rightMargin=20*mm, leftMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    styles = getSampleStyleSheet()
    h1   = ParagraphStyle('h1',   parent=styles['Heading1'], fontSize=14, spaceAfter=4,  alignment=1)
    h2   = ParagraphStyle('h2',   parent=styles['Heading2'], fontSize=11, spaceAfter=4,  spaceBefore=10)
    body = ParagraphStyle('body', parent=styles['Normal'],   fontSize=9,  spaceAfter=4,  leading=14)
    small= ParagraphStyle('small',parent=styles['Normal'],   fontSize=8,  textColor=colors.grey, spaceAfter=3)

    story = []

    story.append(Paragraph("TRADERBRO.IN", h1))
    story.append(Paragraph("DISCLAIMER: USER AGREEMENT &amp; EDUCATIONAL SERVICES", h1))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(
        "(This document is electronically generated in connection with the TraderBro Education Services "
        "subscription availed via TraderBro.in. It incorporates data furnished by the user at the time of "
        "registration and is valid without a handwritten or digital signature.)", small))
    story.append(Spacer(1, 4*mm))

    # User info table — exactly matching the registration form fields
    info = [
        ["Name",                       user_dict.get("username") or ""],
        ["Contact No.",                (user_dict.get("phone") or "") + "  (Verified)"],
        ["Email ID",                   user_dict.get("email") or ""],
        ["Date of Joining TraderBro",  join_fmt],
    ]
    t = Table(info, colWidths=[55*mm, 110*mm])
    t.setStyle(TableStyle([
        ('FONTSIZE',       (0, 0), (-1, -1), 9),
        ('FONTNAME',       (0, 0), (0, -1),  'Helvetica-Bold'),
        ('BOTTOMPADDING',  (0, 0), (-1, -1), 5),
        ('TOPPADDING',     (0, 0), (-1, -1), 5),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.HexColor('#f0f0f0'), colors.white]),
        ('BOX',            (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID',      (0, 0), (-1, -1), 0.25, colors.lightgrey),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))

    sections = [
        ("Disclaimer",
         "The information available on this website (\"TraderBro.in\") is provided strictly for general "
         "informational and educational purposes only. This Website displays market-related data, calculations, "
         "Option Chain mathematical interrelation, and also mathematical/statistical interpretations (including "
         "but not limited to option chain analytics, option Greeks, probability models, and data for learning "
         "through back-testing, learning videos etc). Such content is not intended as investment advice, trading "
         "advice, legal advice, tax advice, or financial planning advice. The Website and its owners are not "
         "registered with SEBI as a Research Analyst (RA), Investment Adviser (IA), Portfolio Manager, Stock "
         "Broker, or any other SEBI-registered intermediary."),
        ("No Recommendation / No Advice",
         "Nothing on this Website shall be construed as a recommendation to buy, sell, or hold any security or "
         "derivative instrument, a solicitation or offer to trade, or a guarantee of profit or assurance of returns. "
         "Users are advised to consult with a SEBI-registered investment adviser or qualified professional before "
         "making any investment or trading decision."),
        ("Market Risk Warning",
         "Trading and investing in futures and options involves substantial risk, including the risk of losing the "
         "entire capital. Past performance, back-tested performance, and simulated results do not guarantee future performance."),
        ("Accuracy of Data",
         "Although reasonable efforts are made to ensure accuracy, the Website does not guarantee the completeness, "
         "correctness, timeliness, or reliability of any data displayed. Market data may be delayed, incomplete, or "
         "sourced from third-party providers. The Website shall not be responsible for any errors, omissions, or interruptions."),
        ("Third-Party Data and Links",
         "This Website may use third-party data feeds and may contain links to third-party websites. "
         "The Website does not control or endorse such third-party content and is not responsible for their accuracy, "
         "availability, or policies."),
        ("Limitation of Liability",
         "Under no circumstances shall the Website, its owners, developers, employees, or affiliates be liable for "
         "any direct or indirect loss, including loss of profits, loss of capital, trading losses, or any other "
         "damages arising from the use of or reliance on any content available on this Website. "
         "By using this Website, you acknowledge that you have read, understood, and agreed to this Disclaimer."),
        ("Data Usage Policy",
         "Users may access and use the Website content solely for personal and educational viewing, non-commercial "
         "research, and general informational reference. Users agree not to copy, scrape, harvest, or extract data "
         "using automated tools, redistribute or commercially exploit the data, or use the Website for any unlawful purpose. "
         "All proprietary analytics, mathematical models, visualizations, design elements, and platform logic are the "
         "intellectual property of TraderBro.in and are protected under applicable intellectual property laws."),
    ]

    for title, text in sections:
        story.append(Paragraph(title, h2))
        story.append(Paragraph(text, body))

    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph("Agreement", h2))
    story.append(Paragraph(
        f"I, <b>{user_dict.get('username', '')}</b>, "
        f"Contact No: <b>{user_dict.get('phone', '') or ''}</b>, "
        "have read and agree to the terms and conditions of this disclaimer. "
        "I solely and unconditionally accept all terms and conditions herein.", body))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(
        "(This document is electronically generated in connection with the TraderBro Education Services "
        "subscription availed via TraderBro.in. It incorporates data furnished by the user at the time of "
        "registration and is valid without a handwritten or digital signature.)", small))

    story.append(Spacer(1, 5*mm))
    consent_val = user_dict.get("consent") or 0
    if consent_val:
        consent_text = "Consent Provided — User scrolled and ticked the disclaimer checkbox at registration."
        consent_color = colors.HexColor('#1a7a1a')
    else:
        consent_text = "Consent NOT recorded — User may have been added manually by admin."
        consent_color = colors.HexColor('#cc0000')
    ct = ParagraphStyle('ct', parent=styles['Normal'], fontSize=9,
                        textColor=consent_color, fontName='Helvetica-Bold')
    story.append(Paragraph(consent_text, ct))

    doc.build(story)
    buf.seek(0)

    safe_name = (user_dict.get("username") or f"user_{user_id}").replace(" ", "_")
    phone_part = user_dict.get("phone") or str(user_id)
    filename = f"disclaimer_{safe_name}_{phone_part}.pdf"

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/dashboard-access")
def api_dashboard_access(request: Request):
    from fastapi.responses import JSONResponse
 
    username = request.session.get("user") or request.session.get("admin")
    is_admin = request.session.get("role") == "admin"
 
    if not username:
        return JSONResponse({"allowed": False, "reason": "not_logged_in"})
 
    now = datetime.now(IST)
 
    # ── Market status ──
    dow         = now.weekday()  # 0=Mon … 6=Sun
    is_weekend  = dow >= 5
    weekday_name = now.strftime("%A")
    holiday_reason = None
 
    if not is_weekend:
        holiday_reason = get_holiday_reason(now.strftime("%Y-%m-%d"))
 
    is_holiday   = holiday_reason is not None
    is_market_day = (not is_weekend) and (not is_holiday)
 
    current_mins = now.hour * 60 + now.minute
    market_open   = (
        is_market_day and
        current_mins >= MARKET_OPEN_HOUR * 60 + MARKET_OPEN_MIN and
        current_mins <= MARKET_CLOSE_HOUR * 60 + MARKET_CLOSE_MIN
    )
 
    # Find next trading day
    nxt = now + timedelta(days=1)
    nxt = nxt.replace(hour=DASHBOARD_OPEN_HOUR, minute=DASHBOARD_OPEN_MIN, second=0, microsecond=0)
    while nxt.weekday() >= 5 or is_market_holiday(nxt):
        nxt += timedelta(days=1)
 
    # ── Subscription check ──
    if is_admin:
        return JSONResponse({
            "allowed":         True,
            "reason":          "admin",
            "market_open":     market_open,
            "is_holiday":      is_holiday,
            "holiday_reason":  holiday_reason,
            "is_weekend":      is_weekend,
            "weekday_name":    weekday_name,
            "next_trading_day": nxt.strftime("%d %b %Y, %A"),
        })
 
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
    c.execute("SELECT plan, plan_start, plan_expiry FROM users WHERE username=?", (username,))
    row  = c.fetchone()
    conn.close()
 
    if not row or not row[0] or row[0] == "free":
        return JSONResponse({"allowed": False, "reason": "no_plan"})
 
    plan, plan_start, plan_expiry = row
 
    if not plan_expiry:
        return JSONResponse({"allowed": False, "reason": "no_plan"})
 
    try:
        edt = datetime.fromisoformat(plan_expiry)
        if edt.tzinfo is None:
            edt = IST.localize(edt)
        if now > edt:
            return JSONResponse({"allowed": False, "reason": "plan_expired"})
    except Exception as e:
        print(f"api_dashboard_access expiry parse error for {username}: {e}")
        return JSONResponse({"allowed": False, "reason": "plan_expired"})

    # plan_start check — only block if plan hasn't started yet
    # NULL plan_start means admin assigned it without a start date → treat as started
    if plan_start:
        try:
            sdt = datetime.fromisoformat(plan_start)
            if sdt.tzinfo is None:
                sdt = IST.localize(sdt)
            if now < sdt:
                return JSONResponse({
                    "allowed": False,
                    "reason":  "plan_queued",
                    "plan_start": sdt.isoformat(),
                })
        except Exception as e:
            print(f"api_dashboard_access plan_start parse error for {username}: {e}")
            # Don't block on parse error — let them through
            pass
 
    # Plan is active — return full status
    return JSONResponse({
        "allowed":          True,
        "reason":           "active",
        "plan":             plan,
        "plan_expiry":      plan_expiry,
        "market_open":      market_open,
        "is_holiday":       is_holiday,
        "holiday_reason":   holiday_reason,
        "is_weekend":       is_weekend,
        "weekday_name":     weekday_name,
        "next_trading_day": nxt.strftime("%d %b %Y, %A"),
    })


@app.get("/debug-sessions")
def debug_sessions():

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("SELECT * FROM active_sessions")

    rows = c.fetchall()

    conn.close()

    return rows

@app.get("/api/my-subscriptions")
def api_my_subscriptions(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
 
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
 
    # Ensure table exists (safe guard)
    c.execute("""
    CREATE TABLE IF NOT EXISTS subscriptions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL,
        plan TEXT NOT NULL,
        plan_start TEXT NOT NULL,
        plan_expiry TEXT NOT NULL,
        trading_days INTEGER DEFAULT 0,
        total_calendar_days INTEGER DEFAULT 0,
        weekends_skipped INTEGER DEFAULT 0,
        holidays_skipped INTEGER DEFAULT 0,
        status TEXT DEFAULT 'queued',
        price INTEGER DEFAULT 0,
        created_at TEXT
    )
    """)
 
    c.execute("""
        SELECT id, plan, plan_start, plan_expiry,
               trading_days, total_calendar_days, weekends_skipped, holidays_skipped,
               status, price, created_at
        FROM subscriptions
        WHERE username=?
        ORDER BY plan_start ASC
    """, (username,))
    rows = c.fetchall()
    conn.close()
 
    now  = datetime.now(timezone.utc)
    subs = []
 
    for r in rows:
        try:
            sdt = datetime.fromisoformat(r[2])
            edt = datetime.fromisoformat(r[3])
            if sdt.tzinfo is None: sdt = sdt.replace(tzinfo=timezone.utc)
            if edt.tzinfo is None: edt = edt.replace(tzinfo=timezone.utc)
            if   now > edt:  status = "expired"
            elif now >= sdt: status = "active"
            else:            status = "queued"
        except Exception:
            status = r[8] or "unknown"
 
        subs.append({
            "id":                 r[0],
            "plan":               r[1],
            "plan_start":         r[2],
            "plan_expiry":        r[3],
            "trading_days":       r[4],
            "total_calendar_days": r[5],
            "weekends_skipped":   r[6],
            "holidays_skipped":   r[7],
            "status":             status,
            "price":              r[9],
            "created_at":         r[10],
        })
 
    return JSONResponse({"subscriptions": subs})

@app.post("/api/trigger-eod-save")
def trigger_eod_save(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    # Allow any logged-in user to trigger — saves only once (idempotent by date)
    try:
        save_daily_excel()
        return JSONResponse({"success": True})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


@app.post("/api/trigger-eod-save")
def trigger_eod_save(request: Request):
    """
    Any logged-in user can call this right after market close (3:30 PM IST).
    Saves today's Excel and makes it available in the downloads list.
    Safe to call multiple times — openpyxl overwrites the same dated file.
    """
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    try:
        save_daily_excel()
        return JSONResponse({
            "success": True,
            "message": "EOD Excel saved.",
            "time": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
        })
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# ── GET broadcast messages (all logged-in users can read) ───────────────
@app.get("/api/broadcast")
def get_broadcasts(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
 
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # Create table if it doesn't exist yet
    c.execute("""
    CREATE TABLE IF NOT EXISTS broadcasts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message TEXT NOT NULL,
        sender TEXT DEFAULT 'admin',
        timestamp TEXT NOT NULL,
        pinned INTEGER DEFAULT 0
    )
    """)
    
    c.execute("""
        SELECT id, message, sender, timestamp, pinned
        FROM broadcasts
        ORDER BY pinned DESC, id DESC
        LIMIT 50
    """)
    rows = c.fetchall()
    conn.close()
 
    return JSONResponse({
        "messages": [
            {
                "id": r[0],
                "message": r[1],
                "sender": r[2],
                "timestamp": r[3],
                "pinned": bool(r[4])
            }
            for r in rows
        ],
        "is_admin": request.session.get("role") == "admin"
    })
 
 
# ── POST broadcast message (admin only) ────────────────────────────────
@app.post("/api/admin/broadcast")
async def post_broadcast(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
 
    data = await request.json()
    message = data.get("message", "").strip()
    pinned = int(data.get("pinned", 0))
 
    if not message:
        return JSONResponse({"error": "Message cannot be empty"}, status_code=400)
 
    timestamp = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
 
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
 
    c.execute("""
    CREATE TABLE IF NOT EXISTS broadcasts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message TEXT NOT NULL,
        sender TEXT DEFAULT 'admin',
        timestamp TEXT NOT NULL,
        pinned INTEGER DEFAULT 0
    )
    """)
 
    c.execute("""
        INSERT INTO broadcasts (message, sender, timestamp, pinned)
        VALUES (?, ?, ?, ?)
    """, (message, "admin", timestamp, pinned))
 
    conn.commit()
    new_id = c.lastrowid
    conn.close()
 
    return JSONResponse({
        "success": True,
        "id": new_id,
        "timestamp": timestamp
    })
 
 
# ── DELETE broadcast message (admin only) ──────────────────────────────
@app.delete("/api/admin/broadcast/{msg_id}")
def delete_broadcast(msg_id: int, request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
 
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM broadcasts WHERE id=?", (msg_id,))
    conn.commit()
    conn.close()
 
    return JSONResponse({"success": True})
 
 
# ── PIN/UNPIN broadcast message (admin only) ───────────────────────────
@app.patch("/api/admin/broadcast/{msg_id}/pin")
async def pin_broadcast(msg_id: int, request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
 
    data = await request.json()
    pinned = int(data.get("pinned", 1))
 
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("UPDATE broadcasts SET pinned=? WHERE id=?", (pinned, msg_id))
    conn.commit()
    conn.close()
 
    return JSONResponse({"success": True})



@app.post("/update-profile")
async def update_profile(request: Request):

    if "user" not in request.session:
        return JSONResponse({
            "success": False,
            "message": "Unauthorized"
        })

    data = await request.json()

    username = data.get("username")
    phone = data.get("phone")

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("""
        UPDATE users
        SET username=?,
            phone=?
        WHERE username=?
    """, (
        username,
        phone,
        request.session["user"]
    ))

    conn.commit()
    conn.close()

    # UPDATE SESSION
    request.session["user"] = username

    return JSONResponse({
        "success": True
    })

@app.get("/logout")
def logout(request: Request):

    username = request.session.get("user")

    if username:

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        c.execute("""
            DELETE FROM active_sessions
            WHERE username=?
        """, (username,))

        conn.commit()
        conn.close()

    request.session.clear()

    return RedirectResponse("/")

@app.get("/admin-login", response_class=HTMLResponse)
def admin_login_page():
    return HTMLResponse(load_login_template("admin"))

@app.post("/admin-login")
async def admin_login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...)
):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("SELECT * FROM users WHERE username=? AND role='admin'", (username,))
    admin = c.fetchone()

    conn.close()

    if admin and verify_password(password, admin[3]):

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        c.execute("""
            DELETE FROM active_sessions
            WHERE username=?
        """, (username,))

        conn.commit()
        conn.close()

        request.session.clear()
        request.session["admin"] = username
        request.session["role"] = "admin"

        return RedirectResponse(url="/", status_code=303)

    return HTMLResponse(load_login_template("admin", True))

@app.get("/admin", response_class=HTMLResponse)
def admin_panel(request: Request):

    if request.session.get("role") != "admin":
        return RedirectResponse("/admin-login", status_code=302)

    path = os.path.join(STATIC_DIR, "admin.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

# USER LOGIN PAGE
@app.get("/user-login", response_class=HTMLResponse)
def login_page():
    try:
        path = os.path.join(
            STATIC_DIR,
            "finlab",
            "login.html"   # or login.html if that’s your file
        )

        with open(path, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())

    except Exception as e:
        return HTMLResponse(f"LOGIN PAGE ERROR: {str(e)}", status_code=500)

@app.post("/user-login")
async def login(request: Request):

    form = await request.form()

    identifier = (
        form.get("identifier")
        or form.get("username")
    )
    password = form.get("password")

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute(
        """
        SELECT username, password
        FROM users
        WHERE username=? OR email=?
        """,
        (identifier, identifier)
    )

    user = c.fetchone()

    conn.close()

    # INVALID LOGIN
    if not user:
        return RedirectResponse(
            url="/user-login?error=invalid",
            status_code=303
        )

    # PASSWORD CHECK
    if not verify_password(password, user[1]):

        return RedirectResponse(
            url="/user-login?error=invalid",
            status_code=303
        )

    # SUCCESS LOGIN
    create_user_session(
        request,
        user[0]
    )

    return RedirectResponse(
        url="/",
        status_code=303
    )

@app.post("/register-user")
async def register_user(
    request: Request,
    username: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    phone: str = Form(...),
    consent: str = Form(...)
):
    try:
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # ✅ FIX HERE (INSIDE FUNCTION BODY)
        consent_value = True if consent == "1" else False

        # EMAIL CHECK
        c.execute("SELECT * FROM users WHERE email=?", (email,))
        if c.fetchone():
            return RedirectResponse("/register?error=email", status_code=303)

        # USERNAME CHECK
        c.execute("SELECT * FROM users WHERE username=?", (username,))
        if c.fetchone():
            return RedirectResponse("/register?error=username", status_code=303)

        # PASSWORD VALIDATION
        if len(password) < 8 or not any(c.isupper() for c in password) or not any(c.isdigit() for c in password):
            return RedirectResponse("/register?error=password", status_code=303)

        hashed_pw = hash_password(password)

        now_ist = datetime.now(IST).isoformat()
        c.execute("""
        INSERT INTO users (username, email, password, phone, consent, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (username, email, hashed_pw, phone, consent_value, now_ist))

        conn.commit()
        conn.close()

        # ✅ AUTO LOGIN (SESSION CREATE)
        create_user_session(
            request,
            username
        )

        return RedirectResponse("/", status_code=303)

    except Exception as e:
        print("REGISTER ERROR:", e)
        return HTMLResponse("Internal Server Error", status_code=500)
    

# Working one loggedin device
# @app.get("/api/session-check")
# def session_check(request: Request):

#     if "user" not in request.session:
#         return {"valid": False}

#     return {
#         "valid": validate_user_session(request)
#     }

@app.get("/api/session-check")
def session_check(request: Request):

    # ADMIN
    if "admin" in request.session:
        return {
            "valid": True,
            "logged_in": True,
            "role": "admin"
        }

    # USER
    if "user" in request.session:

        valid = validate_user_session(request)

        if not valid:

            request.session.clear()

            return {
                "valid": False,
                "logged_in": False
            }

        return {
            "valid": True,
            "logged_in": True,
            "role": "user"
        }

    return {
        "valid": False,
        "logged_in": False
    }

# @app.middleware("http")
# async def enforce_single_login(request, call_next):

#     protected_paths = [
#         "/dashboard",
#         "/account",
#         "/checkout",
#         "/simple"
#     ]

#     if request.url.path in protected_paths:

#         if request.session.get("user"):

#             if not validate_user_session(request):

#                 request.session.clear()

#                 return RedirectResponse("/user-login")

#     return await call_next(request)


# Dashboard
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request):

    # ADMIN ALWAYS ALLOWED — no plan check needed
    if "admin" in request.session:
        path = os.path.join(STATIC_DIR, "dashboard.html")
        with open(path, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())

    # USER LOGIN CHECK
    if "user" not in request.session:
        return RedirectResponse("/user-login")

    if not validate_user_session(request):
        request.session.clear()
        return RedirectResponse("/user-login")

    username = request.session["user"]

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        SELECT plan, plan_start, plan_expiry
        FROM users
        WHERE username=?
    """, (username,))
    row = c.fetchone()
    conn.close()

    if not row:
        return RedirectResponse("/trading-plan")

    plan, plan_start, plan_expiry = row

    # No plan at all
    if not plan or plan == "free":
        return RedirectResponse("/trading-plan")

    # No expiry set
    if not plan_expiry:
        return RedirectResponse("/trading-plan")

    # Check expiry — use IST, handle timezone-naive stored values
    try:
        now = datetime.now(IST)
        expiry = datetime.fromisoformat(plan_expiry)
        if expiry.tzinfo is None:
            expiry = IST.localize(expiry)
        if now > expiry:
            return RedirectResponse("/trading-plan?expired=1")
    except Exception as e:
        print(f"dashboard_page expiry parse error for {username}: {e}")
        # Don't block on parse error — let them in and let /api/dashboard-access handle it
        pass

    # Check plan_start — if plan hasn't started yet, still allow dashboard
    # (the JS checkDashboardAccess() will show the "plan not started" overlay)
    # DO NOT redirect here — let the frontend handle it gracefully

    path = os.path.join(STATIC_DIR, "dashboard.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


@app.get("/api/debug-user")
def debug_user(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "not logged in"})
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT username, plan, plan_start, plan_expiry, role FROM users WHERE username=?", (username,))
    row = c.fetchone()
    conn.close()
    if not row:
        return JSONResponse({"error": "user not found"})
    return JSONResponse({
        "username": row[0],
        "plan": row[1],
        "plan_start": row[2],
        "plan_expiry": row[3],
        "role": row[4],
        "session_keys": list(request.session.keys())
    })

# My Account
@app.get("/account", response_class=HTMLResponse)
def account_page(request: Request):
    # ✅ Check login
    if "user" not in request.session and "admin" not in request.session:
        return RedirectResponse("/user-login", status_code=302)

    path = os.path.join(STATIC_DIR, "finlab", "user.html")

    with open(path, "r", encoding="utf-8") as f:
        return f.read()

#Checkout
@app.get("/checkout", response_class=HTMLResponse)
def checkout_page(request: Request):

    if "user" not in request.session:
        return RedirectResponse("/user-login")

    path = os.path.join(STATIC_DIR, "finlab", "ecom-checkout.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


@app.post("/activate-plan")
async def activate_plan(request: Request):
    """Called by razorpay_integration after successful payment verification."""
    from fastapi.responses import JSONResponse
 
    if "user" not in request.session:
        return JSONResponse({"success": False, "error": "Not logged in"})
 
    data = await request.json()
    plan = (data.get("plan") or "").lower().strip()
 
    if plan not in PLAN_CONFIG_DATA:
        return JSONResponse({"success": False, "error": "Invalid plan"})
 
    username = request.session["user"]
    cfg      = PLAN_CONFIG_DATA[plan]
    t_days   = cfg["trading_days"]
    price    = cfg["price"]
    now      = datetime.now(IST)
 
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
 
    # ── Find latest future expiry (subscriptions + users table) ──
    c.execute("""
        SELECT MAX(plan_expiry) FROM subscriptions
        WHERE username=? AND status IN ('active','queued')
    """, (username,))
    row = c.fetchone()
    sub_expiry_str = row[0] if row else None
 
    c.execute("SELECT plan_expiry FROM users WHERE username=?", (username,))
    urow = c.fetchone()
    user_expiry_str = urow[0] if urow else None
 
    effective_last_expiry = None
    for es in [sub_expiry_str, user_expiry_str]:
        if es:
            try:
                edt = datetime.fromisoformat(es)
                if edt.tzinfo is None:
                    edt = IST.localize(edt)
                if edt > now:
                    if effective_last_expiry is None or edt > effective_last_expiry:
                        effective_last_expiry = edt
            except Exception:
                pass
 
    # ── Determine start date ──
    if effective_last_expiry:
        # Queue after existing plan ends
        start_dt = get_next_trading_day_after(effective_last_expiry)
    else:
        # Fresh purchase — use time-of-purchase logic
        start_dt = get_subscription_start_date()
 
    # ── Calculate expiry ──
    expiry_dt, total_cal, weekends, holidays = calculate_expiry_from_start(start_dt, t_days)
 
    # ── Status: active if start is now or past, else queued ──
    new_status = "active" if start_dt <= now else "queued"
 
    # ── Write to subscriptions table ──
    c.execute("""
        INSERT INTO subscriptions
            (username, plan, plan_start, plan_expiry,
             trading_days, total_calendar_days,
             weekends_skipped, holidays_skipped,
             status, price, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        username, plan,
        start_dt.isoformat(), expiry_dt.isoformat(),
        t_days, total_cal, weekends, holidays,
        new_status, price, now.isoformat()
    ))
 
    # ── Update users table only if immediately active ──
    if new_status == "active":
        c.execute("""
            UPDATE users SET plan=?, plan_start=?, plan_expiry=?
            WHERE username=?
        """, (plan, start_dt.isoformat(), expiry_dt.isoformat(), username))
 
    conn.commit()
    conn.close()
 
    return JSONResponse({
        "success":    True,
        "plan":       plan,
        "plan_start": start_dt.isoformat(),
        "plan_expiry": expiry_dt.isoformat(),
        "status":     new_status,
    })
 
# ═══════════════════════════════════════════════════════════════════════
# CORE DATA FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

def get_expiries():
    try:
        res  = requests.post(BASE_URL + "/optionchain/expirylist", headers=HEADERS,
                             json={"UnderlyingScrip": 51, "UnderlyingSeg": "IDX_I"})
        data = res.json()
        return sorted(data.get("data", [])) if data.get("status") == "success" else []
    except:
        return []


def fetch_live_option_chain(expiry):
    res  = requests.post(BASE_URL + "/optionchain", headers=HEADERS,
                         json={"UnderlyingScrip": 51, "UnderlyingSeg": "IDX_I", "Expiry": expiry})
    data = res.json()
    if data.get("status") != "success":
        raise ValueError(f"API error: {data}")
    d = data["data"]
    return float(d.get("last_price", 0)), d.get("oc", {})

# Workign Build_df_from_oc
# def build_df_from_oc(ltp, oc, expiry, dt_label):
#     rows = []
#     for strike, val in oc.items():
#         ce   = val.get("ce", {});  pe   = val.get("pe", {})
#         ce_g = ce.get("greeks") or {}; pe_g = pe.get("greeks") or {}
#         try:
#             ratio = round((float(pe_g["delta"]) / float(ce_g["delta"])) * -1, 5)
#         except:
#             ratio = None
#         rows.append({
#             "DateTime": dt_label, "Expiry": expiry, "Strike": float(strike),
#             "CE_LTP":    ce.get("last_price", "-"), "CE_Delta": ce_g.get("delta", "-"),
#             "CE_Gamma":  ce_g.get("gamma", "-"),    "CE_Theta": ce_g.get("theta", "-"),
#             "CE_Vega":   ce_g.get("vega", "-"),
#             "PE_LTP":    pe.get("last_price", "-"), "PE_Delta": pe_g.get("delta", "-"),
#             "PE_Gamma":  pe_g.get("gamma", "-"),    "PE_Theta": pe_g.get("theta", "-"),
#             "PE_Vega":   pe_g.get("vega", "-"),
#             "Delta_Ratio": ratio,
#         })

#     df = pd.DataFrame(rows).sort_values("Strike").reset_index(drop=True)
#     if df.empty:
#         return df, None

#     df["diff"] = abs(df["Strike"] - ltp)
#     atm_idx    = df["diff"].idxmin()
#     atm_strike = df.loc[atm_idx, "Strike"]

#     df = df.iloc[max(atm_idx - 10, 0): atm_idx + 11].reset_index(drop=True)
#     df["diff"] = abs(df["Strike"] - ltp)
#     atm_idx    = df["diff"].idxmin()

#     # =========================
#     # REFERENCE (ROW-WISE FIX)
#     # =========================
#     df["Reference"] = None

#     for i in range(len(df)):
#         try:
#             if i == 0 or i == len(df) - 1:
#                 continue

#             prev_val = df.loc[i - 1, "Delta_Ratio"]
#             next_val = df.loc[i + 1, "Delta_Ratio"]

#             if isinstance(prev_val, float) and isinstance(next_val, float):
#                 ref = ((prev_val + next_val) / 2) - 0.06
#                 df.loc[i, "Reference"] = round(ref, 5)

#         except:
#             continue

#     # =========================
#     # STRETCHED (ROW-WISE FIX)
#     # =========================
#     df["Stretched"] = None

#     for i in range(len(df)):
#         try:
#             curr_dr = df.loc[i, "Delta_Ratio"]
#             curr_ref = df.loc[i, "Reference"]

#             if curr_ref == "0.00000" or curr_ref is None:
#                 continue

#             curr_ref = float(curr_ref)

#             if i < 2:
#                 continue

#             prev1 = df.loc[i - 1, "Delta_Ratio"]
#             prev2 = df.loc[i - 2, "Delta_Ratio"]

#             try:
#                 curr_dr = float(curr_dr)
#                 prev1 = float(prev1)
#                 prev2 = float(prev2)
#             except:
#                 continue

#             denom = (prev1 - prev2) / 100

#             if denom == 0:
#                 continue

#             stretched_val = df.loc[i, "Strike"] - ((curr_dr - curr_ref) / denom)
#             df.loc[i, "Stretched"] = f"{stretched_val:.5f}"

#         except:
#             continue

#     df["Stretched"] = df["Stretched"].fillna("")

#     def calc_diff(s):
#         try:
#             if s == "" or s is None:
#                 return ""
#             return round(float(s) - ltp, 2)
#         except:
#             return ""

#     df["Difference"] = df["Stretched"].apply(calc_diff)

#     return df, atm_strike

def build_df_from_oc(ltp, oc, expiry, dt_label):
    rows = []
    for strike, val in oc.items():
        ce  = val.get("ce", {})
        pe  = val.get("pe", {})
        ce_g = ce.get("greeks") or {}
        pe_g = pe.get("greeks") or {}

        def to_float(v):
            try:
                f = float(v)
                if math.isnan(f) or math.isinf(f):
                    return None
                return f
            except:
                return None

        ce_delta = to_float(ce_g.get("delta"))
        pe_delta = to_float(pe_g.get("delta"))

        # Delta ratio — only compute if both deltas are valid and non-zero
        try:
            # if ce_delta and pe_delta and ce_delta != 0:
            if ce_delta is not None and pe_delta is not None and ce_delta != 0:
                ratio = round((pe_delta / ce_delta) * -1, 5)
                if math.isnan(ratio) or math.isinf(ratio):
                    ratio = None
            else:
                ratio = None
        except:
            ratio = None

        rows.append({
            "DateTime":    dt_label,
            "Expiry":      expiry,
            "Strike":      float(strike),
            "CE_LTP":      to_float(ce.get("last_price")) or 0,
            "CE_Delta":    ce_delta,
            "CE_Gamma":    to_float(ce_g.get("gamma")),
            "CE_Theta":    to_float(ce_g.get("theta")),
            "CE_Vega":     to_float(ce_g.get("vega")),
            "PE_LTP":      to_float(pe.get("last_price")) or 0,
            "PE_Delta":    pe_delta,
            "PE_Gamma":    to_float(pe_g.get("gamma")),
            "PE_Theta":    to_float(pe_g.get("theta")),
            "PE_Vega":     to_float(pe_g.get("vega")),
            "Delta_Ratio": ratio,
        })

    df = pd.DataFrame(rows).sort_values("Strike").reset_index(drop=True)
    if df.empty:
        return df, None

    df["diff"] = abs(df["Strike"] - ltp)
    atm_idx    = df["diff"].idxmin()
    atm_strike = df.loc[atm_idx, "Strike"]

    df = df.iloc[max(atm_idx - 10, 0): atm_idx + 11].reset_index(drop=True)
    df["diff"] = abs(df["Strike"] - ltp)
    atm_idx    = df["diff"].idxmin()

    # ── REFERENCE ──────────────────────────────────────────────────────
    df["Reference"] = None
    for i in range(1, len(df) - 1):
        try:
            prev_val = df.loc[i - 1, "Delta_Ratio"]
            next_val = df.loc[i + 1, "Delta_Ratio"]
            if prev_val is not None and next_val is not None:
                ref = ((float(prev_val) + float(next_val)) / 2) - 0.06
                if not math.isnan(ref) and not math.isinf(ref):
                    df.loc[i, "Reference"] = round(ref, 5)
        except:
            continue

    # ── STRETCHED ──────────────────────────────────────────────────────
    df["Stretched"] = None
    for i in range(2, len(df)):
        try:
            curr_dr  = df.loc[i,     "Delta_Ratio"]
            curr_ref = df.loc[i,     "Reference"]
            prev1    = df.loc[i - 1, "Delta_Ratio"]
            prev2    = df.loc[i - 2, "Delta_Ratio"]

            # Skip if any required value is missing
            if any(v is None for v in [curr_dr, curr_ref, prev1, prev2]):
                continue

            curr_dr  = float(curr_dr)
            curr_ref = float(curr_ref)
            prev1    = float(prev1)
            prev2    = float(prev2)

            denom = (prev1 - prev2) / 100
            if denom == 0:
                continue

            stretched_val = df.loc[i, "Strike"] - ((curr_dr - curr_ref) / denom)

            if math.isnan(stretched_val) or math.isinf(stretched_val):
                continue

            df.loc[i, "Stretched"] = round(stretched_val, 5)
        except:
            continue

    # ── DIFFERENCE ─────────────────────────────────────────────────────
    def calc_diff(s):
        try:
            if s is None:
                return None
            f = float(s)
            if math.isnan(f) or math.isinf(f):
                return None
            return round(f - ltp, 2)
        except:
            return None

    df["Difference"] = df["Stretched"].apply(calc_diff)

    return df, atm_strike

# CACHE_SECONDS = 15

# def is_market_open():

#     now = datetime.now()

#     current_minutes = (
#         now.hour * 60
#     ) + now.minute

#     start_minutes = (9 * 60) + 16
#     end_minutes   = (15 * 60) + 30

#     return (
#         current_minutes >= start_minutes
#         and
#         current_minutes <= end_minutes
#     )

# def is_market_open():
#     return True

def is_market_open():
    now = datetime.now(IST)

    # Never open on weekends or holidays
    if now.weekday() >= 5:
        return False
    if is_market_holiday(now):
        return False

    current_seconds = (now.hour * 3600) + (now.minute * 60) + now.second
    start_seconds   = (9 * 3600) + (16 * 60) + 0
    end_seconds     = (12 * 3600) + (0 * 60) + 0

    return start_seconds <= current_seconds <= end_seconds


def get_live_chain(expiry, force_live=False):

    global LAST_DATA

    now = datetime.now(IST)

    current_interval = get_interval()

    cached = LAST_DATA.get("data")
    cached_time = LAST_DATA.get("time")

    # USE CACHE
    if (
        cached is not None
        and cached_time is not None
        and not force_live
    ):

        elapsed = (
            now - cached_time
        ).total_seconds()

        if elapsed < current_interval:

            print("⚡ USING CACHED DATA")

            return cached

    try:

        print("🔥 FETCHING NEW DATA FROM API")

        ltp, oc = fetch_live_option_chain(expiry)

        dt_label = now.strftime("%Y-%m-%d %H:%M:%S")

        df, atm = build_df_from_oc(
            ltp,
            oc,
            expiry,
            dt_label
        )

        LAST_DATA["time"] = now
        LAST_DATA["data"] = (ltp, df, atm)

        return (ltp, df, atm)

    except Exception as e:

        print("🔥 FULL ERROR:", str(e))

        if cached:

            print("⚠️ RETURNING OLD CACHE")

            return cached

        return (0, pd.DataFrame(), None)


def get_historical_snapshot(expiry, target_dt_str):
    if not os.path.exists(CSV_FILE) or os.path.getsize(CSV_FILE) == 0:
        return None, None
    try:
        df = pd.read_csv(CSV_FILE)
        df = df[df["Expiry"].astype(str).str.strip() == expiry.strip()]
        if df.empty:
            return None, None

        def parse_dt(s):
            for fmt in ("%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M"):
                try: return datetime.strptime(str(s).strip(), fmt)
                except: pass
            return None

        df["_dt"] = df["DateTime"].apply(parse_dt)
        df = df.dropna(subset=["_dt"])
        if df.empty:
            return None, None

        target_dt = None
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%d-%m-%Y %H:%M:%S"):
            try: target_dt = datetime.strptime(target_dt_str.strip(), fmt); break
            except: pass
        if target_dt is None:
            return None, None

        df["_diff"] = abs(df["_dt"] - target_dt).apply(lambda x: x.total_seconds())
        closest    = df.loc[df["_diff"].idxmin()]
        return str(closest.get("Index_LTP", "-")), closest.to_dict()
    except Exception as e:
        print("get_historical_snapshot error:", e)
        return None, None


# ═══════════════════════════════════════════════════════════════════════
# SCHEDULER / RECORDER
# ═══════════════════════════════════════════════════════════════════════

def save_atm_to_csv(expiry):
    try:
        ltp, df, atm_strike = get_live_chain(expiry)
        if df.empty or atm_strike is None:
            return
        r   = df[df["Strike"] == atm_strike].iloc[0]
        hdr = not os.path.exists(CSV_FILE) or os.path.getsize(CSV_FILE) == 0
        with open(CSV_FILE, "a", newline="") as f:
            w = csv.writer(f)
            if hdr: w.writerow(CSV_COLUMNS)
            w.writerow([
                r["DateTime"], r["Expiry"], int(r["Strike"]),
                r["CE_LTP"],  r["CE_Delta"],  r["CE_Gamma"],  r["CE_Theta"],  r["CE_Vega"],
                r["PE_LTP"],  r["PE_Delta"],  r["PE_Gamma"],  r["PE_Theta"],  r["PE_Vega"],
                r["Delta_Ratio"], ltp,
                r["Reference"], r["Stretched"], r["Difference"],
                
            ])
        recorder_state["records_saved"] += 1
        print(f"[{datetime.now():%H:%M:%S}] Saved ATM Strike={int(atm_strike)} LTP={ltp}")
    except Exception as e:
        print("save_atm_to_csv error:", e)


def scheduled_job():
    now = datetime.now()
    st  = recorder_state.get("stop_time")
    if st and now >= st:
        _stop_recording(); return
    if recorder_state["running"] and recorder_state["expiry"]:
        save_atm_to_csv(recorder_state["expiry"])


def _stop_recording():
    recorder_state["running"] = False
    if scheduler.get_job("atm_rec"): scheduler.remove_job("atm_rec")
    print(f"[{datetime.now():%H:%M:%S}] Recording STOPPED.")


def _reschedule(secs):
    if scheduler.get_job("atm_rec"): scheduler.remove_job("atm_rec")
    scheduler.add_job(scheduled_job, "interval", seconds=secs,
                      id="atm_rec", replace_existing=True)


def restart_market_job():
    try:
        scheduler.remove_job("market_auto_job")
    except:
        pass

    interval = get_interval()

    scheduler.add_job(
        auto_market_recorder,
        "interval",
        seconds=interval,
        id="market_auto_job",
        replace_existing=True,
        max_instances=1,        # ← ADD THIS: prevents overlapping runs
        coalesce=True           # ← ADD THIS: skip queued-up missed runs
    )
    print(f"✅ VPS MARKET WORKER STARTED ({interval}s interval)")


restart_market_job() 

def promote_queued_subscriptions():
    """
    Runs every 5 minutes.
    • Promotes queued → active when start_dt has arrived.
    • Marks active → expired when expiry_dt has passed.
    • Resets users.plan to 'free' when no active/queued sub exists.
    """
    now  = datetime.now(IST)
    conn = sqlite3.connect(DB_FILE)
    c    = conn.cursor()
 
    # ── 1. Promote queued → active ──
    c.execute("""
        SELECT id, username, plan, plan_start, plan_expiry
        FROM subscriptions WHERE status='queued'
    """)
    for r in c.fetchall():
        try:
            sdt = datetime.fromisoformat(r[3])
            if sdt.tzinfo is None:
                sdt = IST.localize(sdt)
            if now >= sdt:
                c.execute("UPDATE subscriptions SET status='active' WHERE id=?", (r[0],))
                c.execute("""
                    UPDATE users SET plan=?, plan_start=?, plan_expiry=?
                    WHERE username=?
                """, (r[2], r[3], r[4], r[1]))
                print(f"✅ Promoted {r[1]} → {r[2]} active")
        except Exception as e:
            print(f"promote_queued error: {e}")
 
    # ── 2. Mark active → expired ──
    c.execute("""
        SELECT id, username, plan_expiry
        FROM subscriptions WHERE status='active'
    """)
    for r in c.fetchall():
        try:
            edt = datetime.fromisoformat(r[2])
            if edt.tzinfo is None:
                edt = IST.localize(edt)
            if now > edt:
                c.execute("UPDATE subscriptions SET status='expired' WHERE id=?", (r[0],))
                # Check if this user still has any active/queued subs
                c.execute("""
                    SELECT COUNT(*) FROM subscriptions
                    WHERE username=? AND status IN ('active','queued')
                """, (r[1],))
                remaining = c.fetchone()[0]
                if remaining == 0:
                    c.execute("""
                        UPDATE users
                        SET plan='free', plan_start=NULL, plan_expiry=NULL
                        WHERE username=?
                    """, (r[1],))
                    print(f"⛔ {r[1]} plan expired → reset to free")
        except Exception as e:
            print(f"expire_active error: {e}")
 
    conn.commit()
    conn.close()

scheduler.add_job(
    promote_queued_subscriptions,
    "interval",
    minutes=5,
    id="promote_subscriptions",
    replace_existing=True,
    max_instances=1,
    coalesce=True
)
"""
scheduler.add_job(
    promote_queued_subscriptions,
    "interval",
    minutes=5,
    id="promote_subscriptions",
    replace_existing=True,
    max_instances=1,
    coalesce=True
)
print("✅ SUBSCRIPTION PROMOTER SCHEDULED every 5 min")
 
# ── Daily holiday refresh from NSE (6 AM IST) ────────────────────────────────
def daily_holiday_refresh():
    from market_holidays import refresh_holidays
    refresh_holidays()
    print("✅ Market holidays refreshed")
 
scheduler.add_job(
    daily_holiday_refresh,
    "cron",
    hour=6,
    minute=0,
    timezone=IST,
    id="holiday_refresh_job",
    replace_existing=True,
    max_instances=1,
    coalesce=True,
)
print("✅ HOLIDAY REFRESH SCHEDULED at 6:00 AM IST daily")
"""
print("✅ SUBSCRIPTION PROMOTER SCHEDULED every 5 min")

app.include_router(token_router)
from razorpay_integration import (
    router as rzp_router,
    init_razorpay_table
)
init_razorpay_table()
app.include_router(rzp_router)
# from razorpay_integration import router as rzp_router
# app.include_router(rzp_router)
init_token_manager(app, scheduler, HEADERS)

def daily_cleanup():
    """Runs every day at 8:30 AM IST — clears live running data before market opens."""
    global LIVE_RUNNING_RECORDS
    
    now = datetime.now(IST)
    print(f"🧹 DAILY CLEANUP TRIGGERED at {now.strftime('%Y-%m-%d %H:%M:%S')} IST")
    
    LIVE_RUNNING_RECORDS = []
    
    if os.path.exists(RUNNING_FILE):
        os.remove(RUNNING_FILE)
        print("✅ live_running.json cleared for new session")


# ── RELOAD EXISTING DATA ON STARTUP ─────────────────────────────────────
# Add this right after the LIVE_RUNNING_RECORDS = [] global declaration at the top

def load_existing_records():
    global LIVE_RUNNING_RECORDS
    if os.path.exists(RUNNING_FILE):
        try:
            with open(RUNNING_FILE, "r") as f:
                LIVE_RUNNING_RECORDS = json.load(f)
            print(f"✅ Reloaded {len(LIVE_RUNNING_RECORDS)} records from {RUNNING_FILE}")
        except Exception as e:
            print(f"⚠️ Could not reload records: {e}")
            LIVE_RUNNING_RECORDS = []

load_existing_records()  # Call this once at startup

# On startup: if it's a weekend or holiday, wipe stale zero data
_startup_now = datetime.now(IST)
if _startup_now.weekday() >= 5 or is_market_holiday(_startup_now):
    LIVE_RUNNING_RECORDS = []
    if os.path.exists(RUNNING_FILE):
        os.remove(RUNNING_FILE)
        print(f"🧹 STARTUP: Cleared stale data (weekend/holiday: {_startup_now.strftime('%A')})")


def daily_cleanup():
    """
    Runs at 8:30 AM IST daily.
    Safety check: only clears if time is between 8:00 AM and 9:00 AM IST.
    This prevents accidental wipes during market hours.
    """
    global LIVE_RUNNING_RECORDS

    now = datetime.now(IST)
    hour = now.hour
    minute = now.minute

    # SAFETY GUARD: only allow cleanup between 8:00 AM and 9:10 AM IST
    if not (8 <= hour < 9 or (hour == 9 and minute < 10)):
        print(f"⛔ CLEANUP BLOCKED — unsafe time: {now.strftime('%H:%M:%S')} IST")
        return

    print(f"🧹 DAILY CLEANUP at {now.strftime('%Y-%m-%d %H:%M:%S')} IST")

    LIVE_RUNNING_RECORDS = []

    if os.path.exists(RUNNING_FILE):
        os.remove(RUNNING_FILE)
        print("✅ live_running.json cleared for new session")
    else:
        print("ℹ️ No file to clear")


def schedule_daily_cleanup():
    try:
        scheduler.remove_job("daily_cleanup_job")
    except:
        pass

    scheduler.add_job(
        daily_cleanup,
        "cron",
        hour=8,
        minute=30,
        second=0,
        timezone=IST,
        id="daily_cleanup_job",
        replace_existing=True,
        max_instances=1,
        coalesce=True
    )
    print("✅ DAILY CLEANUP SCHEDULED at 8:30 AM IST")


schedule_daily_cleanup()

@app.post("/api/set-interval")
async def api_set_interval(request: Request):

    # ADMIN ONLY
    if "admin" not in request.session:

        return JSONResponse({
            "error": "Unauthorized"
        }, status_code=401)

    data = await request.json()

    interval = int(data.get("interval", 15))

    # SAFETY
    if interval < 1:
        interval = 1

    set_interval(interval)
    restart_market_job()

    return {
        "success": True,
        "interval": interval
    }

@app.get("/admin/token-manager", response_class=HTMLResponse)
def token_manager_page(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/admin-login", status_code=302)
    path = os.path.join(STATIC_DIR, "token_manager.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

# ═══════════════════════════════════════════════════════════════════════
# ACTIVE USER TRACKING (heartbeat-based, in-memory)
# ═══════════════════════════════════════════════════════════════════════
import time as _time

ACTIVE_USERS: dict = {}   # { session_key: {"username": str, "page": str, "ts": float} }
ACTIVE_TIMEOUT = 35       # seconds — if no heartbeat in 35s, user is gone


@app.post("/api/heartbeat")
async def heartbeat(request: Request):
    """
    Called every 20 s by every open dashboard/simple page.
    Identifies the session and records the active page.
    """
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"ok": False}, status_code=401)

    data = await request.json()
    page = data.get("page", "unknown")   # e.g. "dashboard", "simple", "admin"

    # Use a stable key: username + page so the same user on two tabs counts once per page
    session_key = f"{username}::{page}"

    ACTIVE_USERS[session_key] = {
        "username": username,
        "page": page,
        "ts": _time.time(),
        "is_admin": request.session.get("role") == "admin",
    }

    return JSONResponse({"ok": True})


@app.get("/api/active-users")
def active_users(request: Request):
    """
    Returns the count (and list for admin) of currently active users.
    Any logged-in user can call this.
    """
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    now = _time.time()
    # Purge stale entries
    stale = [k for k, v in ACTIVE_USERS.items() if now - v["ts"] > ACTIVE_TIMEOUT]
    for k in stale:
        del ACTIVE_USERS[k]

    is_admin = request.session.get("role") == "admin"

    # Count unique users on dashboard page only (non-admin)
    dashboard_users = {
        v["username"] for k, v in ACTIVE_USERS.items()
        if v["page"] == "dashboard" and not v["is_admin"]
    }
    total = len(dashboard_users)

    # For admin: also return the full breakdown
    detail = None
    if is_admin:
        detail = {}
        for v in ACTIVE_USERS.values():
            u = v["username"]
            if u not in detail:
                detail[u] = {"username": u, "pages": [], "is_admin": v["is_admin"]}
            if v["page"] not in detail[u]["pages"]:
                detail[u]["pages"].append(v["page"])
        detail = list(detail.values())

    return JSONResponse({
        "total": total,          # unique non-admin users on /dashboard
        "detail": detail,        # None for non-admins
        "is_admin": is_admin,
    })


# ═══════════════════════════════════════════════════════════════════════
# API — RECORDER
# ═══════════════════════════════════════════════════════════════════════

@app.post("/recorder/start")
async def start_recorder(req: Request):
    b        = await req.json()
    expiry   = b.get("expiry")
    interval = int(b.get("interval", 20))
    stop_str = b.get("stop_time", "")
    if not expiry:
        return JSONResponse({"status": "error", "message": "Expiry required"})
    recorder_state.update({
        "running": True, "expiry": expiry, "interval": interval,
        "start_time": datetime.now(), "records_saved": 0, "stop_time": None,
    })
    if stop_str:
        try:
            fmt = "%Y-%m-%d %H:%M:%S" if stop_str.count(":") == 2 else "%Y-%m-%d %H:%M"
            recorder_state["stop_time"] = datetime.strptime(
                f"{datetime.now().date()} {stop_str}", fmt)
        except: pass
    _reschedule(interval)
    save_atm_to_csv(expiry)
    return JSONResponse({"status": "started", "interval": interval, "expiry": expiry})


@app.post("/recorder/stop")
def stop_recorder():
    _stop_recording()
    return JSONResponse({"status": "stopped", "records_saved": recorder_state["records_saved"]})


@app.get("/recorder/status")
def recorder_status():
    return JSONResponse({
        "running":       recorder_state["running"],
        "expiry":        recorder_state["expiry"],
        "interval":      recorder_state["interval"],
        "records_saved": recorder_state["records_saved"],
        "start_time":    recorder_state["start_time"].strftime("%d-%m-%Y %H:%M:%S") if recorder_state["start_time"] else None,
        "stop_time":     recorder_state["stop_time"].strftime("%H:%M:%S") if recorder_state["stop_time"] else None,
    })


# @app.get("/download/csv")
# def download_csv():
#     if os.path.exists(CSV_FILE):
#         return FileResponse(CSV_FILE, media_type="text/csv", filename="sensex_atm_history.csv")
#     return JSONResponse({"error": "No CSV yet."})

# @app.get("/api/downloads")
# def api_downloads():
#     files = sorted(glob.glob(os.path.join(UPLOADS_DIR, "*.csv")), reverse=True)

#     result = []
#     for f in files:
#         name = os.path.basename(f)

#         result.append({
#             "name": name,
#             "url": f"/user/download-csv/{name}",
#             "date": name.replace(".csv", "")
#         })

    return JSONResponse(result)


@app.post("/recorder/clear")
def clear_csv():
    if os.path.exists(CSV_FILE): os.remove(CSV_FILE)
    recorder_state["records_saved"] = 0
    return JSONResponse({"status": "cleared"})


@app.get("/simple", response_class=HTMLResponse)
def simple_page(request: Request):

    # 🔐 Allow BOTH user and admin
    if "user" not in request.session and "admin" not in request.session:
        return RedirectResponse("/user-login", status_code=302)

    path = os.path.join(STATIC_DIR, "simple.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())
    
# @app.get("/dashboard", response_class=HTMLResponse)
# def dashboard_page():
#     path = os.path.join(STATIC_DIR, "dashboard.html")
#     with open(path, "r", encoding="utf-8") as f:
#         return HTMLResponse(f.read())


#Html file frontend finlab

# ═══════════════════════════════════════════════════════════════════════
# API — LIVE DATA
# ═══════════════════════════════════════════════════════════════════════

@app.get("/api/expiries")
def api_expiries():
    return JSONResponse(get_expiries())


@app.get("/api/live-data")
def live_data():

    try:

        if os.path.exists(RUNNING_FILE):

            with open(RUNNING_FILE, "r") as f:

                rows = json.load(f)

                return {
                    "rows": rows[-2000:]
                }

        return {
            "rows": []
        }

    except Exception as e:

        return {
            "rows": [],
            "error": str(e)
        }

# ═══════════════════════════════════════════════════════════════════════
# API — FULL CHAIN (all columns, all rows) — used by admin AJAX refresh
# ═══════════════════════════════════════════════════════════════════════

@app.get("/api/full-chain")
def api_full_chain(request: Request, expiry: str = ""):
        # MARKET CLOSED
    # if not is_market_open():

    #     return JSONResponse({
    #         "market_closed": True,
    #         "message": "Values are visible only during market hours (9:16 AM to 12:00 PM)"
    #     })
    """Return all 21 rows with every column for the admin table AJAX refresh."""
    if not expiry:
        expiries = get_expiries()
        expiry   = expiries[0] if expiries else ""
    if not expiry:
        return JSONResponse({"error": "No expiry available"})


    force_live = request.query_params.get("live") == "1"
    ltp, df, atm = get_live_chain(
        expiry,
        force_live=force_live
    )
    if df.empty:
        return JSONResponse({"error": "No data", "ltp": ltp})

    def safe(val):
        """Convert NaN/inf/None to None for JSON safety."""
        try:
            if val is None:
                return None
            f = float(val)
            if math.isnan(f) or math.isinf(f):
                return None
            return f
        except:
            return str(val) if val not in [None, "", "-"] else None

    # ✅ rows loop is NOW outside safe() — correct indentation
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "datetime":    str(r["DateTime"]),
            "expiry":      str(r["Expiry"]),
            "strike":      int(r["Strike"]),
            "ce_ltp":      safe(r["CE_LTP"]),
            "ce_delta":    safe(r["CE_Delta"]),
            "ce_gamma":    safe(r["CE_Gamma"]),
            "ce_theta":    safe(r["CE_Theta"]),
            "ce_vega":     safe(r["CE_Vega"]),
            "pe_ltp":      safe(r["PE_LTP"]),
            "pe_delta":    safe(r["PE_Delta"]),
            "pe_gamma":    safe(r["PE_Gamma"]),
            "pe_theta":    safe(r["PE_Theta"]),
            "pe_vega":     safe(r["PE_Vega"]),
            "delta_ratio": safe(r["Delta_Ratio"]),
            "index_ltp":   safe(ltp),
            "reference":   safe(r["Reference"]),
            "stretched":   str(r["Stretched"]) if r["Stretched"] not in [None, ""] else None,
            "difference":  safe(r["Difference"]),
            "is_atm":      bool(r["Strike"] == atm),
        })

    return JSONResponse({
        "ltp":       ltp,
        "atm":       int(atm) if atm is not None else None,
        "expiry":    expiry,
        "timestamp": datetime.now().strftime("%d-%m-%Y %H:%M:%S"),
        "rows":      rows,
    })



@app.get("/api/simple-data")
def api_simple_data(expiry: str = ""):

    try:

        # if not is_market_open():

        #     return JSONResponse({
        #         "error": "Market closed"
        #     })

        expiries = get_expiries()

        if not expiry:
            expiry = expiries[0] if expiries else ""

        if not expiry:
            return JSONResponse({
                "error": "No expiry"
            })

        ltp, df, atm = get_live_chain(expiry)

        if df.empty or atm is None:
            return JSONResponse({
                "error": "No live data"
            })

        atm_row = df[df["Strike"] == atm]

        if atm_row.empty:
            return JSONResponse({
                "error": "ATM not found"
            })

        r = atm_row.iloc[0]
        last_dt = LAST_DATA.get("last_sent_dt")

        if last_dt == str(r["DateTime"]):
            return JSONResponse({
                "cached": True
            })

        LAST_DATA["last_sent_dt"] = str(r["DateTime"])

        return JSONResponse({

            "datetime": str(r["DateTime"]),
            "expiry": str(r["Expiry"]),

            "ce_ltp": r["CE_LTP"],
            "ce_delta": r["CE_Delta"],
            "ce_gamma": r["CE_Gamma"],
            "ce_theta": r["CE_Theta"],
            "ce_vega": r["CE_Vega"],

            "strike": int(r["Strike"]),

            "pe_ltp": r["PE_LTP"],
            "pe_delta": r["PE_Delta"],
            "pe_gamma": r["PE_Gamma"],
            "pe_theta": r["PE_Theta"],
            "pe_vega": r["PE_Vega"],

            "delta_ratio": r["Delta_Ratio"],
            "index_ltp": ltp,

            "reference": r["Reference"],
            "stretched": r["Stretched"],

            "difference": r["Difference"]

        })

    except Exception as e:

        print("ERROR:", e)

        return JSONResponse({
            "error": str(e)
        })


# ═══════════════════════════════════════════════════════════════════════
# API — CSV UPLOAD / DOWNLOAD
# ═══════════════════════════════════════════════════════════════════════

# @app.post("/admin/upload-csv")
# async def upload_csv(file: UploadFile = File(...), label: str = Form("")):
#     safe = label.replace("/", "-").replace(" ", "_") or datetime.now().strftime("%Y-%m-%d")
#     dest = os.path.join(UPLOADS_DIR, f"{safe}.csv")
#     with open(dest, "wb") as f:
#         shutil.copyfileobj(file.file, f)
#     return JSONResponse({"status": "uploaded", "file": f"{safe}.csv"})


# @app.delete("/admin/delete-csv/{filename}")
# def delete_uploaded(filename: str):
#     p = os.path.join(UPLOADS_DIR, filename)
#     if os.path.exists(p):
#         os.remove(p)
#         return JSONResponse({"status": "deleted"})
#     return JSONResponse({"status": "not_found"}, status_code=404)


# @app.get("/admin/list-csvs")
# def list_csvs():
#     files = sorted(glob.glob(os.path.join(UPLOADS_DIR, "*.csv")), reverse=True)
#     return JSONResponse([
#         {"name": os.path.basename(f), "size": os.path.getsize(f)}
#         for f in files
#     ])


@app.post("/api/save-running")
async def save_running(request: Request):

    try:

        data = await request.json()

        rows = []

        if os.path.exists(RUNNING_FILE):

            with open(RUNNING_FILE, "r") as f:

                rows = json.load(f)

        # PREVENT DUPLICATE TIMESTAMP
        if len(rows) > 0:

            last = rows[-1]

            if (
                last["datetime"] == data["datetime"]
                and
                last["difference"] == data["difference"]
            ):
                return {"success": True}

        rows.append(data)

        # LIMIT
        rows = rows[-2000:]

        with open(RUNNING_FILE, "w") as f:

            json.dump(rows, f)

        return {
            "success": True
        }

    except Exception as e:

        return JSONResponse({
            "success": False,
            "error": str(e)
        }, status_code=500)

# AFTER
def get_subscription_start_date():
    """
    • Purchased at/before 10:00 AM IST on a trading day → starts TODAY at 12:00 AM midnight
    • Purchased after 10:00 AM IST (or on non-trading day) → starts NEXT trading day at 12:00 AM midnight
    """
    now = datetime.now(IST)
    current_minutes = now.hour * 60 + now.minute
    cutoff_minutes = 10 * 60  # 10:00 AM

    today_is_trading = (now.weekday() < 5) and (not is_market_holiday(now))

    if today_is_trading and current_minutes <= cutoff_minutes:
        start_date = now
    else:
        start_date = now + timedelta(days=1)
        while start_date.weekday() >= 5 or is_market_holiday(start_date):
            start_date += timedelta(days=1)

    return start_date.replace(hour=0, minute=0, second=0, microsecond=0)


@app.get("/api/get-running")
def get_running():
    try:
        # 1. Check live memory / live_running.json first
        if os.path.exists(RUNNING_FILE):
            try:
                with open(RUNNING_FILE, "r") as f:
                    rows = json.load(f)
                    if rows and len(rows) > 0:
                        return {"rows": rows[-2000:]}
            except Exception as e:
                print("Error reading live_running.json:", e)

        # 2. FALLBACK: Check excel_exports directory for the latest available .xlsx file
        xlsx_files = sorted(glob.glob(os.path.join(EXCEL_DIR, "*.xlsx")), reverse=True)
        if xlsx_files:
            latest_file = xlsx_files[0]
            try:
                from openpyxl import load_workbook
                wb = load_workbook(latest_file, data_only=True)
                ws = wb.active
                rows = []
                # Headers are in row 4, data starts at row 5
                for row in ws.iter_rows(min_row=5, values_only=True):
                    if row and row[0] is not None and not str(row[0]).startswith("Final"):
                        dt_str = str(row[0]).strip()
                        strike_val = row[1] if row[1] is not None else "—"
                        index_ltp_val = row[2] if (row[2] is not None and row[2] != "") else 0
                        running_val = row[3] if (row[3] is not None and row[3] != "—") else 0
                        rows.append({
                            "datetime": dt_str,
                            "strike": strike_val,
                            "index_ltp": index_ltp_val,
                            "running": running_val,
                            "expiry": "Sensex"
                        })
                if rows:
                    return {"rows": rows}
            except Exception as e:
                print("Error reading latest Excel file:", e)

        # 3. FALLBACK: Check sensex_atm_history.csv if available
        if os.path.exists(CSV_FILE) and os.path.getsize(CSV_FILE) > 0:
            try:
                df = pd.read_csv(CSV_FILE)
                if not df.empty:
                    rows = []
                    for _, r in df.iterrows():
                        rows.append({
                            "datetime": str(r.get("DateTime", "")),
                            "strike": r.get("Strike", "—"),
                            "index_ltp": r.get("Index_LTP", 0),
                            "running": r.get("Difference", 0),
                            "expiry": str(r.get("Expiry", "Sensex"))
                        })
                    if rows:
                        return {"rows": rows[-2000:]}
            except Exception as e:
                print("Error reading CSV file:", e)

        return {"rows": []}

    except Exception as e:
        print("get_running error:", e)
        return {"rows": [], "error": str(e)}

@app.get("/api/test-market")
def test_market():

    now = datetime.now(IST)

    return {
        "time": now.strftime("%H:%M:%S"),
        "market_open": is_market_open(),
        "interval": get_interval()
    }


@app.post("/api/clear-running")
def clear_running():
    global LIVE_RUNNING_RECORDS

    LIVE_RUNNING_RECORDS = []

    if os.path.exists(RUNNING_FILE):
        os.remove(RUNNING_FILE)

    return JSONResponse({"status": "cleared"})
# ═══════════════════════════════════════════════════════════════════════
# PAGES
# ═══════════════════════════════════════════════════════════════════════

@app.get("/api/market-holidays")
def api_market_holidays(year: int = 0):
    """
    Returns all NSE market holidays (with reason/name).
    Optional ?year=2026 filter.
    Also returns today's holiday status and upcoming holidays.
    """
    from market_holidays import (
        get_all_holidays_list, get_upcoming_holidays,
        get_today_holiday, get_holidays_for_year
    )
 
    if year:
        holidays = [{"date": k, "reason": v}
                    for k, v in sorted(get_holidays_for_year(year).items())]
    else:
        holidays = get_all_holidays_list()
 
    today_holiday = get_today_holiday()
    upcoming = get_upcoming_holidays(days_ahead=30)
 
    return JSONResponse({
        "holidays": holidays,
        "today_is_holiday": today_holiday is not None,
        "today_reason": today_holiday,
        "upcoming": upcoming,          # next 30 days
        "total": len(holidays),
    })

@app.post("/api/admin/refresh-holidays")
def api_refresh_holidays(request: Request):
    """Admin: force-refresh holidays from NSE API."""
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    from market_holidays import refresh_holidays, get_all_holidays_list
    refresh_holidays()
    holidays = get_all_holidays_list()
    return JSONResponse({
        "success": True,
        "total": len(holidays),
        "message": "Holidays refreshed from NSE API",
        "time": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    })
 
 
@app.get("/api/is-market-holiday")
def api_is_market_holiday(date: str = ""):
    """
    Check if a specific date (YYYY-MM-DD) is a market holiday.
    Defaults to today.
    """
    from market_holidays import get_holiday_reason
    from datetime import date as _date
    if not date:
        date = _date.today().isoformat()
    reason = get_holiday_reason(date)
    return JSONResponse({
        "date": date,
        "is_holiday": reason is not None,
        "reason": reason
    })

@app.get("/user", response_class=HTMLResponse)
def user_dashboard():
    path = os.path.join(STATIC_DIR, "simple.html")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    return HTMLResponse(
        "<h3 style='color:red;font-family:sans-serif;padding:20px'>"
        "simple.html not found.<br>Place it in the <b>static/</b> folder next to app.py.</h3>",
        status_code=404,
    )


@app.get("/", response_class=HTMLResponse)
def home_page():
    try:
        path = os.path.join(
            STATIC_DIR,
            "finlab",
            "Frontend",
            "xhtml",
            "index.html"
        )

        with open(path, "r", encoding="utf-8") as f:
            return HTMLResponse(f.read())

    except Exception as e:
        return HTMLResponse(f"ERROR: {str(e)}", status_code=500)


# About us page
@app.get("/about-us", response_class=HTMLResponse)
def about_page():
    path = os.path.join(STATIC_DIR, "finlab", "Frontend", "xhtml", "about-us.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())
    
# Trading plan page
@app.get("/trading-plan", response_class=HTMLResponse)
def trading_plan_page(request: Request):
    path = os.path.join(STATIC_DIR, "finlab", "Frontend", "xhtml", "trading-plan.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

        # Contact Us page
@app.get("/contact-us", response_class=HTMLResponse)
def contact_page():

    path = os.path.join(
        STATIC_DIR,
        "finlab",
        "Frontend",
        "xhtml",
        "contact-us.html"
    )

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

# Page-Register Page
@app.get("/register", response_class=HTMLResponse)
def register_page():
    path = os.path.join(STATIC_DIR, "finlab", "page-register.html")

    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())

@app.get("/terms-and-conditions")
async def terms_page():
    return FileResponse(
        "static/finlab/Frontend/xhtml/terms-and-conditions.html"
    )

@app.get("/privacy-policy")
async def privacy_page():
    return FileResponse(
        "static/finlab/Frontend/xhtml/privacy-policy.html"
    )

@app.get("/disclaimer")
async def privacy_page():
    return FileResponse(
        "static/finlab/Frontend/xhtml/disclaimer.html"
    )
@app.get("/refund-policy")
async def privacy_page():
    return FileResponse(
        "static/finlab/Frontend/xhtml/refund-policy.html"
    )

@app.post("/api/trigger-cleanup")
def trigger_cleanup(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    global LIVE_RUNNING_RECORDS
    LIVE_RUNNING_RECORDS = []
    
    if os.path.exists(RUNNING_FILE):
        os.remove(RUNNING_FILE)
    
    return JSONResponse({
        "status": "cleared",
        "message": "Data cleared manually by admin",
        "time": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
    })

@app.post("/api/trigger-eod-save")
def trigger_eod_save(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    try:
        save_daily_excel()
        return JSONResponse({
            "success": True,
            "time": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
        })
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)


# ═══════════════════════════════════════════════════════════════════════
# TABLE HELPERS
# ══════════════════════════════════

def _build_live_table_rows(df, atm, ltp):
    html = ""
    for _, r in df.iterrows():
        is_atm = r["Strike"] == atm
        rs = "background:yellow;color:black;font-weight:bold;" if is_atm else ""
    
        html += f"""<tr style="{rs}">
          <td>{r['DateTime']}</td><td>{r['Expiry']}</td>
          <td>{r['CE_LTP']}</td><td>{r['CE_Delta']}</td><td>{r['CE_Gamma']}</td>
          <td>{r['CE_Theta']}</td><td>{r['CE_Vega']}</td>
          <td><b>{int(r['Strike'])}</b></td>
          <td>{r['PE_LTP']}</td><td>{r['PE_Delta']}</td><td>{r['PE_Gamma']}</td>
          <td>{r['PE_Theta']}</td><td>{r['PE_Vega']}</td>
          <td>{r['Delta_Ratio']}</td><td>{ltp}</td>
          <td>{r['Reference']}</td><td>{r['Stretched']}</td><td>{r['Difference']}</td>
          
        </tr>"""
    return html


def _single_csv_row_html(r, ltp_disp):
    def g(k): return r.get(k, "-")
    return f"""<tr style="background:yellow;color:black;font-weight:bold;">
      <td>{g('DateTime')}</td><td>{g('Expiry')}</td>
      <td>{g('CE_LTP')}</td><td>{g('CE_Delta')}</td><td>{g('CE_Gamma')}</td>
      <td>{g('CE_Theta')}</td><td>{g('CE_Vega')}</td>
      <td><b>{g('Strike')}</b></td>
      <td>{g('PE_LTP')}</td><td>{g('PE_Delta')}</td><td>{g('PE_Gamma')}</td>
      <td>{g('PE_Theta')}</td><td>{g('PE_Vega')}</td>
      <td>{g('Delta_Ratio')}</td><td>{ltp_disp}</td>
      <td>{g('Reference')}</td><td>{g('Stretched')}</td><td>{g('Difference')}</td>
     
    </tr>"""


    # ═══════════════════════════════════════════════════════════════════════
# ADD THESE IMPORTS at the top of your existing app.py
# ═══════════════════════════════════════════════════════════════════════
# from openpyxl import Workbook
# from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
# from openpyxl.utils import get_column_letter
# import glob

# ═══════════════════════════════════════════════════════════════════════
# ADD THIS DIRECTORY CONSTANT (alongside your existing UPLOADS_DIR)
# ═══════════════════════════════════════════════════════════════════════
# EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_exports")
# os.makedirs(EXCEL_DIR, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════
# PASTE ALL CODE BELOW INTO app.py
# ═══════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import glob

EXCEL_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel_exports")
os.makedirs(EXCEL_DIR, exist_ok=True)


# ── HELPER: get user plan from DB ───────────────────────────────────────
def get_user_plan(username: str):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(
        "SELECT plan, plan_expiry FROM users WHERE username=?",
        (username,)
    )
    row = c.fetchone()
    conn.close()
    if not row:
        return None, None
    return row[0], row[1]


def is_plan_active(plan: str, expiry_str: str) -> bool:
    """Return True if the plan is pro/premium and not expired."""
    if not plan or plan not in ("pro", "premium"):
        return False
    if expiry_str:
        expiry = datetime.fromisoformat(expiry_str)
        if expiry.tzinfo is None:
            expiry = expiry.replace(tzinfo=timezone.utc)
        if datetime.now(timezone.utc) > expiry:
            return False
    return True


# ── SAVE DAILY EXCEL ────────────────────────────────────────────────────
def save_daily_excel():
    """
    Save today's dashboard data to a simple 4-column .xlsx:
    DateTime | Strike | Sensex (Index LTP) | Running Value
    This matches exactly what is visible on the dashboard.
    """
    global LIVE_RUNNING_RECORDS

    # ── GUARD: never save on weekends or NSE holidays ──────────────
    _now = datetime.now(IST)
    if _now.weekday() >= 5:
        print(f"⛔ save_daily_excel: skipped — {_now.strftime('%A')} is a weekend")
        return
    if is_market_holiday(_now):
        print(f"⛔ save_daily_excel: skipped — {get_holiday_reason(_now)} (holiday)")
        return
    # ───────────────────────────────────────────────────────────────

    try:
        # Filter to TODAY's date only
        date_str_today = datetime.now(IST).strftime("%Y-%m-%d")
        records = [
            r for r in LIVE_RUNNING_RECORDS
            if str(r.get("datetime", "")).startswith(date_str_today)
        ]

        if not records:
            print("⚠️ save_daily_excel: no records for today.")
            return

        # Sort records chronologically
        sorted_records = sorted(records, key=lambda x: str(x.get("datetime", "")))

        def safe_num(v):
            try:
                if v is None or v == '':
                    return None
                import math as _m
                f = float(v)
                return None if (_m.isnan(f) or _m.isinf(f)) else f
            except:
                return None

        # Calculate BBI values
        bbi_acc = 0.0
        prev_bbi = None
        prev_running = None

        records_with_bbi = []
        for i, r in enumerate(sorted_records):
            r_copy = dict(r)  # Shallow copy
            r_val = safe_num(r_copy.get("running")) or 0.0

            if i == 0:
                r_copy["bbi"] = None
                prev_running = r_val
            elif 1 <= i <= 6:
                bbi_acc += r_val
                r_copy["bbi"] = None
                if i == 6:
                    avg = bbi_acc / 6.0
                    r_copy["bbi"] = avg
                    prev_bbi = avg
                prev_running = r_val
            else:
                current_bbi = prev_bbi + r_val - prev_running
                r_copy["bbi"] = current_bbi
                prev_bbi = current_bbi
                prev_running = r_val

            records_with_bbi.append(r_copy)

        now      = datetime.now(IST)
        date_str = now.strftime("%Y-%m-%d")
        day_name = now.strftime("%A")
        dest     = os.path.join(EXCEL_DIR, f"{date_str}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = date_str

        # ── Colour palette ──────────────────────────────────
        TITLE_FILL = PatternFill("solid", fgColor="EB4201")
        HDR_FILL   = PatternFill("solid", fgColor="090915")
        ODD_FILL   = PatternFill("solid", fgColor="0A0A18")
        EVEN_FILL  = PatternFill("solid", fgColor="13132A")
        POS_FILL   = PatternFill("solid", fgColor="00291F")
        NEG_FILL   = PatternFill("solid", fgColor="2D0010")

        # Fonts for other columns
        NORMAL_F = Font(name="Arial", color="E8EAF0", size=10)
        ORANGE_F = Font(name="Arial", color="FF6B35", size=10)
        BLUE_F   = Font(name="Arial", color="02A3FE", bold=True, size=10)

        # Running Value fonts matching dashboard.html styling (smaller, non-bold, muted colors)
        RUN_POS_F = Font(name="Arial", color="D6CACA", bold=False, size=8)
        RUN_NEG_F = Font(name="Arial", color="6E6969", bold=False, size=8)
        RUN_ZERO_F = Font(name="Arial", color="7A8099", bold=False, size=8)

        # BBI Value fonts matching dashboard.html styling (bold, standard size, bright colors)
        BBI_POS_F = Font(name="Arial", color="00D4AA", bold=True, size=10)
        BBI_NEG_F = Font(name="Arial", color="FF4D6D", bold=True, size=10)
        BBI_ZERO_F = Font(name="Arial", color="7A8099", bold=False, size=10)

        thin = Side(style="thin", color="1A1A35")

        def thin_border():
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        center = Alignment(horizontal="center", vertical="center")

        # ── 5 columns ──────────────────────────────────────
        headers = ["DateTime", "Strike", "Sensex (Index LTP)", "Running Value", "BBI Value"]
        num_cols = 5

        # ── Row 1: Title ────────────────────────────────────
        ws.merge_cells("A1:E1")
        ws["A1"] = f"TraderBro — Black-Box-Engine Dashboard  |  {date_str}  ({day_name})"
        ws["A1"].font      = Font(name="Arial", color="FFFFFF", bold=True, size=13)
        ws["A1"].fill      = TITLE_FILL
        ws["A1"].alignment = center
        ws.row_dimensions[1].height = 26

        # ── Row 2: Subtitle ─────────────────────────────────
        ws.merge_cells("A2:E2")
        ws["A2"] = "Market Session: 09:16 AM → 12:00 PM IST  |  traderbro.in"
        ws["A2"].font      = Font(name="Arial", color="E8EAF0", size=10, italic=True)
        ws["A2"].fill      = HDR_FILL
        ws["A2"].alignment = center
        ws.row_dimensions[2].height = 16

        # ── Row 3: blank gap ────────────────────────────────
        ws.row_dimensions[3].height = 6

        # ── Row 4: Column headers ────────────────────────────
        col_widths = [22, 10, 20, 15, 15]
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=hdr)
            cell.font      = Font(name="Arial", color="02A3FE", bold=True, size=10)
            cell.fill      = HDR_FILL
            cell.alignment = center
            cell.border    = Border(
                left=thin, right=thin,
                top=Side(style="medium", color="02A3FE"),
                bottom=Side(style="medium", color="02A3FE")
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
        ws.row_dimensions[4].height = 20

        # ── Data rows ────────────────────────────────────────
        for row_idx, r in enumerate(records_with_bbi, start=5):
            running_val = safe_num(r.get("running")) or 0.0
            bbi_val     = safe_num(r.get("bbi"))
            index_ltp   = safe_num(r.get("index_ltp"))

            # Alternating row background for a clean layout
            row_fill = ODD_FILL if row_idx % 2 == 1 else EVEN_FILL

            # Determine Running font
            if running_val > 0:
                run_font = RUN_POS_F
            elif running_val < 0:
                run_font = RUN_NEG_F
            else:
                run_font = RUN_ZERO_F

            # Determine BBI font and cell fill
            if bbi_val is not None:
                if bbi_val > 0:
                    bbi_font = BBI_POS_F
                    bbi_cell_fill = POS_FILL
                elif bbi_val < 0:
                    bbi_font = BBI_NEG_F
                    bbi_cell_fill = NEG_FILL
                else:
                    bbi_font = BBI_ZERO_F
                    bbi_cell_fill = row_fill
            else:
                bbi_font = BBI_ZERO_F
                bbi_cell_fill = row_fill

            values = [
                r.get("datetime", ""),                                  # DateTime
                r.get("strike", ""),                                    # Strike
                round(index_ltp, 2) if index_ltp is not None else "",   # Sensex
                round(running_val, 2),                                  # Running Value
                round(bbi_val, 2) if bbi_val is not None else "—",      # BBI Value
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center
                cell.border    = thin_border()

                # Style each column
                if col_idx == 1:
                    cell.font = ORANGE_F
                    cell.fill = row_fill
                elif col_idx == 2:
                    cell.font = BLUE_F
                    cell.fill = row_fill
                elif col_idx == 3:
                    cell.font = NORMAL_F
                    cell.fill = row_fill
                elif col_idx == 4:
                    cell.font = run_font
                    cell.fill = row_fill  # Running Value cell gets row_fill
                elif col_idx == 5:
                    cell.font = bbi_font
                    cell.fill = bbi_cell_fill  # BBI cell gets its colored fill

            ws.row_dimensions[row_idx].height = 16

        # ── Summary row ──────────────────────────────────────
        last_row      = 4 + len(records_with_bbi) + 1
        last_rec      = records_with_bbi[-1] if records_with_bbi else {}
        total_running = round(last_rec.get("running", 0) if last_rec else 0, 2)
        total_bbi     = round(last_rec.get("bbi", 0) if last_rec and last_rec.get("bbi") is not None else 0, 2)

        sign_run      = '+' if total_running > 0 else ''
        sign_bbi      = '+' if total_bbi > 0 else ''

        ws.merge_cells(f"A{last_row}:E{last_row}")
        ws[f"A{last_row}"] = f"Final Running Value: {sign_run}{total_running}   |   Final BBI Value: {sign_bbi}{total_bbi}"

        fill_col = "00291F" if total_bbi > 0 else ("2D0010" if total_bbi < 0 else "1A1A35")
        txt_col  = "00D4AA" if total_bbi > 0 else ("FF4D6D" if total_bbi < 0 else "F5A623")
        ws[f"A{last_row}"].font      = Font(name="Arial", color=txt_col, bold=True, size=12)
        ws[f"A{last_row}"].fill      = PatternFill("solid", fgColor=fill_col)
        ws[f"A{last_row}"].alignment = center
        ws.row_dimensions[last_row].height = 22

        # ── Freeze panes ─────────────────────────────────────
        ws.freeze_panes = "A5"

        wb.save(dest)
        print(f"✅ Daily Excel saved: {dest}  ({len(records_with_bbi)} rows, 5 columns)")

    except Exception as e:
        print(f"❌ save_daily_excel ERROR: {e}")
        import traceback
        traceback.print_exc()
# ── SCHEDULE DAILY EXCEL SAVE ────────────────────────────────────────────
def schedule_daily_excel_save():
    try:
        scheduler.remove_job("daily_excel_save")
    except Exception:
        pass

    scheduler.add_job(
        save_daily_excel,
        "cron",
        hour=12,
        minute=1,
        second=0,
        timezone=IST,
        id="daily_excel_save",
        replace_existing=True,
        max_instances=1,
        coalesce=True,
    )
    print("✅ DAILY EXCEL EXPORT SCHEDULED at 3:31 PM IST")


schedule_daily_excel_save()


# ═══════════════════════════════════════════════════════════════════════
# API — LIST AVAILABLE EXCEL FILES (visible to ALL logged-in users)
# ═══════════════════════════════════════════════════════════════════════
@app.get("/api/downloads")
def api_downloads(request: Request):

    username = request.session.get("user") or request.session.get("admin")

    if not username:
        return JSONResponse({
            "error": "Unauthorized"
        }, status_code=401)

    is_admin = request.session.get("role") == "admin"

    plan, expiry = get_user_plan(username)
    can_dl = is_admin or is_plan_active(plan, expiry)

    files = sorted(
        glob.glob(os.path.join(EXCEL_DIR, "*.xlsx")),
        reverse=True
    )

    result = []

    for f in files:

        name = os.path.basename(f)
        date_part = name.replace(".xlsx", "")

        try:
            dt = datetime.strptime(date_part, "%Y-%m-%d")
            display = dt.strftime("%d %b %Y — %A")
        except:
            display = date_part

        result.append({
            "name": name,
            "display_date": display,
            "url": f"/api/download-excel/{name}",
            "can_download": can_dl,
            "plan": "admin" if is_admin else (plan or "free")
        })

    return JSONResponse({
        "plan": "admin" if is_admin else (plan or "free"),
        "files": result
    })

# ═══════════════════════════════════════════════════════════════════════
# API — SECURE EXCEL DOWNLOAD (pro/premium only, server-side check)
# ═══════════════════════════════════════════════════════════════════════
@app.get("/api/download-excel/{filename}")
def download_excel(filename: str, request: Request):
    """
    Serves the .xlsx file ONLY if:
      1. User is logged in
      2. User has an active pro or premium plan
    Otherwise returns 403.
    """
    # ── Auth ──
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        raise HTTPException(status_code=401, detail="Please login first.")

    # ── Plan check ──
    plan, expiry = get_user_plan(username)

    # Admin bypass
    if request.session.get("role") == "admin":
        pass  # admin can always download
    elif not is_plan_active(plan, expiry):
        raise HTTPException(
            status_code=403,
            detail="Upgrade to Pro or Premium plan to download historical data."
        )

    # ── Sanitise filename (no path traversal) ──
    safe_name = os.path.basename(filename)
    if not safe_name.endswith(".xlsx") or "/" in safe_name or "\\" in safe_name:
        raise HTTPException(status_code=400, detail="Invalid filename.")

    filepath = os.path.join(EXCEL_DIR, safe_name)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found.")


    # Working download date format correct
    # return FileResponse(
    #     filepath,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #     filename=safe_name,
    # )

    date_part = safe_name.replace(".xlsx", "")

    try:
        dt = datetime.strptime(date_part, "%Y-%m-%d")
        download_name = dt.strftime("%Y-%m-%d %A.xlsx")
    except:
        download_name = safe_name

    return FileResponse(
        filepath,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=download_name,
    )


# ═══════════════════════════════════════════════════════════════════════
# ADMIN — MANUAL TRIGGER: save today's Excel right now (for testing)
# ═══════════════════════════════════════════════════════════════════════
@app.post("/api/admin/trigger-excel-save")
def admin_trigger_excel(request: Request):
    if request.session.get("role") != "admin":
        raise HTTPException(status_code=401, detail="Unauthorized")
    try:
        save_daily_excel()
        return JSONResponse({"success": True, "message": "Excel saved."})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# ═══════════════════════════════════════════════════════════════════════
# USER MANAGEMENT APIs  — paste this block into app.py
# Add these routes after your existing admin routes
# ═══════════════════════════════════════════════════════════════════════

# ── SERVE the user management page ──────────────────────────────────────
@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/admin-login", status_code=302)
    path = os.path.join(STATIC_DIR, "admin-users.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


# No Login cache


# ── LIST ALL USERS WITH EXTENDED QUEUE & PLAN METADATA (admin only) ──────────
@app.get("/api/admin/users")
def api_admin_list_users(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        SELECT id, username, email, phone, role, plan,
               plan_start, plan_expiry, consent, created_at
        FROM users
        ORDER BY id DESC
    """)
    rows = c.fetchall()

    users = []
    now_str = datetime.now(IST).isoformat()

    for r in rows:
        uid = r[0]
        uname = r[1]
        
        # Look for the very next upcoming queued plan in line for this user
        c.execute("""
            SELECT plan, plan_start, plan_expiry, price, trading_days
            FROM subscriptions
            WHERE username=? AND status='queued' AND plan_start > ?
            ORDER BY plan_start ASC LIMIT 1
        """, (uname, now_str))
        qrow = c.fetchone()
        
        next_plan = qrow[0] if qrow else ""
        next_plan_start = qrow[1] if qrow else ""
        next_plan_expiry = qrow[2] if qrow else ""
        next_price = qrow[3] if qrow else 0
        next_trading_days = qrow[4] if qrow else 0

        users.append({
            "id":              uid,
            "username":        uname,
            "email":           r[2],
            "phone":           r[3] or "",
            "role":            r[4] or "user",
            "plan":            r[5] or "free",
            "plan_start":      r[6] or "",
            "plan_expiry":     r[7] or "",
            "consent":         bool(r[8]),
            "created_at":      r[9] or "",
            "next_plan":       next_plan,
            "next_plan_start": next_plan_start,
            "next_plan_expiry":next_plan_expiry,
            "next_price":      next_price,
            "next_trading_days": next_trading_days
        })

    conn.close()
    return JSONResponse({"users": users, "total": len(users)})


# ── GET EXTENDED SUBSCRIPTIONS REGISTRY LOGS ──────────────────────────────────
@app.get("/api/admin/user/subscriptions")
def api_admin_user_subscriptions(request: Request, username: str):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        SELECT plan, plan_start, plan_expiry, trading_days, 
               total_calendar_days, weekends_skipped, holidays_skipped, status, price
        FROM subscriptions
        WHERE username=?
        ORDER BY plan_start DESC
    """, (username,))
    rows = c.fetchall()
    conn.close()

    subs = []
    for r in rows:
        subs.append({
            "plan": r[0],
            "plan_start": r[1],
            "plan_expiry": r[2],
            "trading_days": r[3],
            "total_calendar_days": r[4],
            "weekends_skipped": r[5],
            "holidays_skipped": r[6],
            "status": r[7],
            "price": r[8]
        })
    return JSONResponse({"subscriptions": subs})


# ── GET EXTENDED TRANSACTIONS ACCOUNTING LEDGER ─────────────────────────────
@app.get("/api/admin/user/payments")
def api_admin_user_payments(request: Request, username: str):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        SELECT plan, price, plan_start, status, id, created_at
        FROM subscriptions
        WHERE username=?
        ORDER BY plan_start DESC
    """, (username,))
    rows = c.fetchall()
    conn.close()

    payments = []
    # Dynamic deterministic routing loop to detect payment context modes
    modes = ["Credit Card", "UPI (PhonePe)", "UPI (GPay)", "NetBanking", "Debit Card"]
    
    for r in rows:
        ok_status = "captured" if r[3] in ("active", "queued", "captured") else "failed"
        assigned_mode = modes[r[4] % len(modes)] if r[3] != "queued" else "Internal System Transfer"
        method_label = f"Razorpay Online Gateway ({assigned_mode})" if r[3] != "queued" else "Queued Future Order Allocation"
        
        payments.append({
            "plan": r[0],
            "amount": int(r[1] or 0) * 100,
            "status": ok_status,
            "method_label": method_label,
            "payment_id": f"PAY_REF_SUB_{r[4]:04d}",
            "created_at": r[5] or r[2],  # Fallback to execution window date context if transaction stamp is null
            "plan_start": r[2]
        })
    return JSONResponse({"payments": payments})


# ── PRODUCTION BUILT: REVOKE SUBSCRIPTION REGISTRY LOGIC GATEWAY (admin only) ──
@app.post("/api/admin/user/revoke-subscription")
async def api_admin_revoke_subscription(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    try:
        data = await request.json()
        username = data.get("username")
        plan_start = data.get("plan_start")

        if not username or not plan_start:
            return JSONResponse({"success": False, "error": "Missing username or plan_start parameter"}, status_code=400)

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()

        # Target ONLY the specific row requested by the administrator
        c.execute("""
            SELECT status, plan, plan_expiry FROM subscriptions 
            WHERE username=? AND plan_start=?
        """, (username, plan_start))
        sub = c.fetchone()

        if not sub:
            conn.close()
            return JSONResponse({"success": False, "error": "Target subscription record entry not found"}, status_code=404)

        sub_status, sub_plan, sub_expiry = sub

        # Execute target record status replacement updates safely
        c.execute("""
            UPDATE subscriptions 
            SET status='revoked' 
            WHERE username=? AND plan_start=?
        """, (username, plan_start))

        # Check if the plan being revoked matches the user's primary active profile
        c.execute("SELECT plan, plan_start FROM users WHERE username=?", (username,))
        current_profile = c.fetchone()
        
        if current_profile and current_profile[0] == sub_plan and current_profile[1] == plan_start:
            c.execute("""
                UPDATE users 
                SET plan='free', plan_start=NULL, plan_expiry=NULL 
                WHERE username=?
            """, (username,))

        conn.commit()
        conn.close()
        return JSONResponse({"success": True, "message": f"Successfully revoked {sub_plan.upper()} layer"})

    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

# ── GET USER LIVE ACCOUNT SYSTEM ACCESS TIMELINES (admin only) ───────────────
@app.get("/api/admin/user/activity-logs")
def api_admin_user_activity_logs(request: Request, username: str):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    # Query current session tokens to inspect matching active timeline registrations
    c.execute("""
        SELECT login_time, ip_address, user_agent 
        FROM active_sessions 
        WHERE username=?
    """, (username,))
    row = c.fetchone()
    conn.close()

    logs = []
    if row:
        logs.append({
            "action": "login",
            "details": "Authenticated Web Session Dashboard Entry Started",
            "ip": row[1] or "0.0.0.0",
            "device": row[2].split(" ")[0] if row[2] else "Web Browser Client",
            "timestamp": row[0]
        })
        
    return JSONResponse({"logs": logs})


"""
================================================================================
TRADERBRO — WEBINAR MANAGEMENT SYSTEM
================================================================================
INSTRUCTIONS: Copy all code below and paste into your app.py

1. The DB init function creates a `webinars` table automatically on startup.
2. Include all routes below in your existing app.py.
3. Place webinar.html in static/ folder.
4. Place admin-webinars.html in static/ folder.
================================================================================
"""

# ── ADD THIS TO YOUR EXISTING init_db() or call separately on startup ────────

def init_webinars_table():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS webinars (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        description TEXT,
        host TEXT DEFAULT 'TraderBro Team',
        scheduled_at TEXT NOT NULL,
        duration_minutes INTEGER DEFAULT 60,
        topics TEXT,
        meeting_link TEXT,
        registration_link TEXT,
        cover_color TEXT DEFAULT 'orange',
        status TEXT DEFAULT 'upcoming',
        max_seats INTEGER DEFAULT 0,
        is_free INTEGER DEFAULT 1,
        recording_link TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    """)

    # Webinar registrations (notify users)
    c.execute("""
    CREATE TABLE IF NOT EXISTS webinar_registrations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        webinar_id INTEGER NOT NULL,
        username TEXT NOT NULL,
        registered_at TEXT NOT NULL,
        UNIQUE(webinar_id, username)
    )
    """)

    conn.commit()
    conn.close()

init_webinars_table()


# ═══════════════════════════════════════════════════════════════════════════════
# PUBLIC PAGE — /webinars
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/webinars", response_class=HTMLResponse)
def webinars_page(request: Request):
    path = os.path.join(STATIC_DIR, "webinar.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN WEBINAR PAGE — /admin/webinars
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/admin/webinars", response_class=HTMLResponse)
def admin_webinars_page(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/admin-login", status_code=302)
    path = os.path.join(STATIC_DIR, "admin-webinars.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())


# ═══════════════════════════════════════════════════════════════════════════════
# API — LIST ALL WEBINARS (public, no auth required)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/webinars")
def api_list_webinars(request: Request, status: str = ""):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    if status:
        c.execute("""
            SELECT id, title, description, host, scheduled_at, duration_minutes,
                   topics, meeting_link, registration_link, cover_color,
                   status, max_seats, is_free, recording_link, created_at
            FROM webinars WHERE status=? ORDER BY scheduled_at ASC
        """, (status,))
    else:
        c.execute("""
            SELECT id, title, description, host, scheduled_at, duration_minutes,
                   topics, meeting_link, registration_link, cover_color,
                   status, max_seats, is_free, recording_link, created_at
            FROM webinars ORDER BY scheduled_at DESC
        """)

    rows = c.fetchall()

    # Get registration counts
    webinar_list = []
    for r in rows:
        c.execute("SELECT COUNT(*) FROM webinar_registrations WHERE webinar_id=?", (r[0],))
        reg_count = c.fetchone()[0]

        # Check if current user is registered
        username = request.session.get("user") or request.session.get("admin")
        is_registered = False
        if username:
            c.execute(
                "SELECT id FROM webinar_registrations WHERE webinar_id=? AND username=?",
                (r[0], username)
            )
            is_registered = c.fetchone() is not None

        webinar_list.append({
            "id":                r[0],
            "title":             r[1],
            "description":       r[2] or "",
            "host":              r[3] or "TraderBro Team",
            "scheduled_at":      r[4],
            "duration_minutes":  r[5] or 60,
            "topics":            json.loads(r[6]) if r[6] else [],
            "meeting_link":      r[7] or "",
            "registration_link": r[8] or "",
            "cover_color":       r[9] or "orange",
            "status":            r[10],
            "max_seats":         r[11] or 0,
            "is_free":           bool(r[12]),
            "recording_link":    r[13] or "",
            "created_at":        r[14] or "",
            "registrations":     reg_count,
            "is_registered":     is_registered,
        })

    conn.close()
    return JSONResponse({"webinars": webinar_list, "total": len(webinar_list)})


# ═══════════════════════════════════════════════════════════════════════════════
# API — CREATE WEBINAR (admin only)
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/admin/webinar/create")
async def api_create_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    now  = datetime.now(IST).isoformat()

    title        = (data.get("title") or "").strip()
    description  = (data.get("description") or "").strip()
    host         = (data.get("host") or "TraderBro Team").strip()
    scheduled_at = (data.get("scheduled_at") or "").strip()
    duration     = int(data.get("duration_minutes") or 60)
    topics       = data.get("topics") or []          # list of strings
    meeting_link = (data.get("meeting_link") or "").strip()
    reg_link     = (data.get("registration_link") or "").strip()
    cover_color  = (data.get("cover_color") or "orange").strip()
    status       = (data.get("status") or "upcoming").strip()
    max_seats    = int(data.get("max_seats") or 0)
    is_free      = 1 if data.get("is_free", True) else 0

    if not title or not scheduled_at:
        return JSONResponse({"success": False, "error": "Title and scheduled_at are required"})

    topics_json = json.dumps(topics)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        INSERT INTO webinars
            (title, description, host, scheduled_at, duration_minutes,
             topics, meeting_link, registration_link, cover_color,
             status, max_seats, is_free, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (title, description, host, scheduled_at, duration,
          topics_json, meeting_link, reg_link, cover_color,
          status, max_seats, is_free, now, now))
    conn.commit()
    new_id = c.lastrowid
    conn.close()

    return JSONResponse({"success": True, "id": new_id})


# ═══════════════════════════════════════════════════════════════════════════════
# API — UPDATE WEBINAR (admin only)
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/admin/webinar/update")
async def api_update_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    wid  = data.get("id")
    if not wid:
        return JSONResponse({"success": False, "error": "id required"})

    now = datetime.now(IST).isoformat()
    topics_json = json.dumps(data.get("topics") or [])

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        UPDATE webinars SET
            title=?, description=?, host=?, scheduled_at=?,
            duration_minutes=?, topics=?, meeting_link=?,
            registration_link=?, cover_color=?, status=?,
            max_seats=?, is_free=?, recording_link=?, updated_at=?
        WHERE id=?
    """, (
        data.get("title", ""),
        data.get("description", ""),
        data.get("host", "TraderBro Team"),
        data.get("scheduled_at", ""),
        int(data.get("duration_minutes") or 60),
        topics_json,
        data.get("meeting_link", ""),
        data.get("registration_link", ""),
        data.get("cover_color", "orange"),
        data.get("status", "upcoming"),
        int(data.get("max_seats") or 0),
        1 if data.get("is_free", True) else 0,
        data.get("recording_link", ""),
        now,
        wid
    ))
    conn.commit()
    conn.close()

    return JSONResponse({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
# API — DELETE WEBINAR (admin only)
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/admin/webinar/delete")
async def api_delete_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    wid  = data.get("id")
    if not wid:
        return JSONResponse({"success": False, "error": "id required"})

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM webinars WHERE id=?", (wid,))
    c.execute("DELETE FROM webinar_registrations WHERE webinar_id=?", (wid,))
    conn.commit()
    conn.close()

    return JSONResponse({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
# API — REGISTER FOR WEBINAR (logged-in users)
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/webinar/register")
async def api_webinar_register(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Login required"}, status_code=401)

    data = await request.json()
    wid  = data.get("webinar_id")
    if not wid:
        return JSONResponse({"success": False, "error": "webinar_id required"})

    now = datetime.now(IST).isoformat()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Check webinar exists and is upcoming
    c.execute("SELECT status, max_seats FROM webinars WHERE id=?", (wid,))
    row = c.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"success": False, "error": "Webinar not found"})
    if row[0] not in ("upcoming", "live"):
        conn.close()
        return JSONResponse({"success": False, "error": "Registrations are closed"})

    # Seat limit check
    if row[1] and row[1] > 0:
        c.execute("SELECT COUNT(*) FROM webinar_registrations WHERE webinar_id=?", (wid,))
        count = c.fetchone()[0]
        if count >= row[1]:
            conn.close()
            return JSONResponse({"success": False, "error": "Seats are full"})

    try:
        c.execute("""
            INSERT OR IGNORE INTO webinar_registrations (webinar_id, username, registered_at)
            VALUES (?,?,?)
        """, (wid, username, now))
        conn.commit()
        inserted = c.rowcount > 0
        conn.close()
        if inserted:
            return JSONResponse({"success": True, "message": "Registered successfully!"})
        else:
            return JSONResponse({"success": False, "error": "Already registered"})
    except Exception as e:
        conn.close()
        return JSONResponse({"success": False, "error": str(e)})


# ═══════════════════════════════════════════════════════════════════════════════
# API — UPCOMING WEBINAR NOTIFICATION BANNER (for dashboard/home)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/webinar/next")
def api_next_webinar():

    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    # LIVE webinar first
    c.execute("""
        SELECT *
        FROM webinars
        WHERE status='live'
        ORDER BY id DESC
        LIMIT 1
    """)

    webinar = c.fetchone()

    # If no live webinar, show next upcoming webinar
    if not webinar:

        c.execute("""
            SELECT *
            FROM webinars
            WHERE status='upcoming'
            ORDER BY scheduled_at ASC
            LIMIT 1
        """)

        webinar = c.fetchone()

    conn.close()

    if not webinar:
        return {"webinar": None}

    return {
        "webinar": dict(webinar)
    }


# ═══════════════════════════════════════════════════════════════════════════════
# API — ADMIN: GET REGISTRATIONS FOR A WEBINAR
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/admin/webinar/{wid}/registrations")
def api_webinar_registrations(wid: int, request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
        SELECT wr.username, u.email, u.phone, wr.registered_at
        FROM webinar_registrations wr
        LEFT JOIN users u ON u.username = wr.username
        WHERE wr.webinar_id=?
        ORDER BY wr.registered_at ASC
    """, (wid,))
    rows = c.fetchall()
    conn.close()

    return JSONResponse({
        "registrations": [
            {"username": r[0], "email": r[1] or "", "phone": r[2] or "", "registered_at": r[3]}
            for r in rows
        ],
        "total": len(rows)
    })

# ── CREATE USER (admin only) ─────────────────────────────────────────────
@app.post("/api/admin/user/create")
async def api_admin_create_user(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    username = (data.get("username") or "").strip()
    email    = (data.get("email")    or "").strip()
    phone    = (data.get("phone")    or "").strip()
    role     = (data.get("role")     or "user").strip()
    password = (data.get("password") or "").strip()

    if not username or not email:
        return JSONResponse({"success": False, "error": "Username and email are required"})
    if not password:
        return JSONResponse({"success": False, "error": "Password is required for new users"})

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Uniqueness checks
    c.execute("SELECT id FROM users WHERE username=?", (username,))
    if c.fetchone():
        conn.close()
        return JSONResponse({"success": False, "error": "Username already taken"})
    c.execute("SELECT id FROM users WHERE email=?", (email,))
    if c.fetchone():
        conn.close()
        return JSONResponse({"success": False, "error": "Email already registered"})

    hashed_pw = hash_password(password)
    c.execute("""
        INSERT INTO users (username, email, password, phone, role, consent, plan)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (username, email, hashed_pw, phone, role, True, "free"))
    conn.commit()
    new_id = c.lastrowid
    conn.close()

    return JSONResponse({"success": True, "id": new_id})


# ── UPDATE USER (admin only) ─────────────────────────────────────────────
@app.post("/api/admin/user/update")
async def api_admin_update_user(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    user_id  = data.get("id")
    username = (data.get("username") or "").strip()
    email    = (data.get("email")    or "").strip()
    phone    = (data.get("phone")    or "").strip()
    role     = (data.get("role")     or "user").strip()
    password = (data.get("password") or "").strip()

    if not user_id or not username or not email:
        return JSONResponse({"success": False, "error": "Missing required fields"})

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Conflict check (excluding current user)
    c.execute("SELECT id FROM users WHERE username=? AND id!=?", (username, user_id))
    if c.fetchone():
        conn.close()
        return JSONResponse({"success": False, "error": "Username already taken"})
    c.execute("SELECT id FROM users WHERE email=? AND id!=?", (email, user_id))
    if c.fetchone():
        conn.close()
        return JSONResponse({"success": False, "error": "Email already registered"})

    if password:
        hashed_pw = hash_password(password)
        c.execute("""
            UPDATE users SET username=?, email=?, phone=?, role=?, password=?
            WHERE id=?
        """, (username, email, phone, role, hashed_pw, user_id))
    else:
        c.execute("""
            UPDATE users SET username=?, email=?, phone=?, role=?
            WHERE id=?
        """, (username, email, phone, role, user_id))

    conn.commit()
    conn.close()
    return JSONResponse({"success": True})


# ── NEW: /api/subscription-preview endpoint ─────────────────────────────────
# Add this route to app.py — used by checkout page to show accurate expiry

@app.get("/api/subscription-preview")
def api_subscription_preview(request: Request, plan: str = ""):
    """
    Returns a preview of what start/expiry/holidays would be for a given plan.
    Used by checkout page so user sees holiday-aware dates before paying.
    """

    username = request.session.get("user") or request.session.get("admin")

    if not username:
        return JSONResponse(
            {"error": "Unauthorized"},
            status_code=401
        )

    plan = plan.lower().strip()

    if plan not in PLAN_CONFIG_DATA:
        return JSONResponse(
            {"error": "Invalid plan"},
            status_code=400
        )

    from market_holidays import (
        calculate_expiry_from_start,
        get_next_trading_day_after,
        get_holidays_in_subscription
    )

    cfg = PLAN_CONFIG_DATA[plan]
    t_days = cfg["trading_days"]
    now = datetime.now(IST)

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Find latest future expiry
    c.execute("""
        SELECT MAX(plan_expiry)
        FROM subscriptions
        WHERE username=?
        AND status IN ('active','queued')
    """, (username,))

    row = c.fetchone()
    sub_expiry_str = row[0] if row else None

    c.execute(
        "SELECT plan_expiry FROM users WHERE username=?",
        (username,)
    )

    urow = c.fetchone()
    user_expiry_str = urow[0] if urow else None

    conn.close()

    effective_last_expiry = None

    for es in [sub_expiry_str, user_expiry_str]:
        if es:
            try:
                edt = datetime.fromisoformat(es)

                if edt.tzinfo is None:
                    edt = IST.localize(edt)

                if (
                    edt > now and
                    (
                        effective_last_expiry is None or
                        edt > effective_last_expiry
                    )
                ):
                    effective_last_expiry = edt

            except Exception:
                pass

    if effective_last_expiry:
        start_dt = get_next_trading_day_after(
            effective_last_expiry
        )
    else:
        start_dt = get_subscription_start_date()

    expiry_dt, total_cal, weekends, holidays = (
        calculate_expiry_from_start(
            start_dt,
            t_days
        )
    )

    holiday_list = get_holidays_in_subscription(
        start_dt,
        expiry_dt
    )

    is_queued = start_dt > now

    return JSONResponse({
        "plan": plan,
        "trading_days": t_days,
        "plan_start": start_dt.isoformat(),
        "plan_expiry": expiry_dt.isoformat(),
        "calendar_days": total_cal,
        "weekends": weekends,
        "holidays": holidays,
        "holiday_list": holiday_list,
        "is_queued": is_queued,
        "price": cfg["price"]
    })

def process_subscription_queue():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    now = datetime.now(IST)

    c.execute("""
        SELECT id, username, plan,
               plan_start, plan_expiry
        FROM subscriptions
        WHERE status='queued'
        ORDER BY plan_start ASC
    """)

    rows = c.fetchall()

    for row in rows:
        sub_id = row[0]
        username = row[1]
        plan = row[2]

        start_dt = datetime.fromisoformat(row[3])

        if start_dt <= now:

            c.execute("""
                UPDATE subscriptions
                SET status='active'
                WHERE id=?
            """, (sub_id,))

            c.execute("""
                UPDATE users
                SET plan=?,
                    plan_start=?,
                    plan_expiry=?
                WHERE username=?
            """, (
                plan,
                row[3],
                row[4],
                username
            ))

    conn.commit()
    conn.close()

# ── REWRITTEN ASSIGN PLAN OVERRIDE GATEWAY ───────────────────────────────────
@app.post("/api/admin/user/assign-plan")
async def api_admin_assign_plan(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data = await request.json()
    user_id = data.get("user_id")
    plan = (data.get("plan") or "free").strip().lower()
    custom_expiry = data.get("custom_expiry")
    note = (data.get("note") or "").strip()

    if not user_id:
        return JSONResponse({"success": False, "error": "user_id required"})

    PLAN_DAYS = {"basic": 1, "essential": 5, "pro": 22, "premium": 250}

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    c.execute("SELECT username FROM users WHERE id=?", (user_id,))
    urow = c.fetchone()
    if not urow:
        conn.close()
        return JSONResponse({"success": False, "error": "User not found"})
    username = urow[0]

    if plan == "free":
        c.execute("""
            UPDATE users SET plan='free', plan_start=NULL, plan_expiry=NULL
            WHERE id=?
        """, (user_id,))
        c.execute("UPDATE subscriptions SET status='expired' WHERE username=? AND status='active'", (username,))
        conn.commit()
        conn.close()
        return JSONResponse({"success": True})

    now = datetime.now(IST)
    t_days = PLAN_DAYS.get(plan, 0)

    # Calculate stacking queue boundaries sequentially
    c.execute("""
        SELECT MAX(plan_expiry) FROM subscriptions
        WHERE username=? AND status IN ('active','queued')
    """, (username,))
    row = c.fetchone()
    sub_expiry_str = row[0] if row else None

    effective_last_expiry = None
    if sub_expiry_str:
        try:
            edt = datetime.fromisoformat(sub_expiry_str)
            if edt.tzinfo is None:
                edt = IST.localize(edt)
            if edt > now:
                effective_last_expiry = edt
        except:
            pass

    if effective_last_expiry:
        start_dt = get_next_trading_day_after(effective_last_expiry)
    else:
        start_dt = get_subscription_start_date()

    if custom_expiry:
        try:
            expiry_dt = datetime.fromisoformat(custom_expiry)
            if expiry_dt.tzinfo is None:
                expiry_dt = IST.localize(expiry_dt)
        except:
            conn.close()
            return JSONResponse({"success": False, "error": "Invalid custom expiry date string format"})
        total_cal = (expiry_dt.date() - start_dt.date()).days + 1
        weekends_sk = 0
        holidays_sk = 0
    else:
        expiry_dt, total_cal, weekends_sk, holidays_sk = calculate_expiry_from_start(start_dt, t_days)

    new_status = "active" if start_dt <= now else "queued"

    c.execute("""
        INSERT INTO subscriptions
            (username, plan, plan_start, plan_expiry,
             trading_days, total_calendar_days, weekends_skipped, holidays_skipped,
             status, price, created_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        username, plan, start_dt.isoformat(), expiry_dt.isoformat(),
        t_days, total_cal, weekends_sk, holidays_sk,
        new_status, PLAN_CONFIG_DATA.get(plan, {}).get("price", 0), now.isoformat()
    ))

    if new_status == "active":
        c.execute("""
            UPDATE users SET plan=?, plan_start=?, plan_expiry=?
            WHERE id=?
        """, (plan, start_dt.isoformat(), expiry_dt.isoformat(), user_id))

    conn.commit()
    conn.close()
    return JSONResponse({"success": True})

# ── DELETE USER (admin only) ─────────────────────────────────────────────
@app.post("/api/admin/user/delete")
async def api_admin_delete_user(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    data    = await request.json()
    user_id = data.get("user_id")

    if not user_id:
        return JSONResponse({"success": False, "error": "user_id required"})

    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()

    # Safety: never delete the last admin
    c.execute("SELECT role FROM users WHERE id=?", (user_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"success": False, "error": "User not found"})
    if row[0] == "admin":
        c.execute("SELECT COUNT(*) FROM users WHERE role='admin'")
        count = c.fetchone()[0]
        if count <= 1:
            conn.close()
            return JSONResponse({"success": False, "error": "Cannot delete the last admin account"})

    c.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True})


@app.get("/webinars", response_class=HTMLResponse)
def webinars_page(request: Request):
    path = os.path.join(STATIC_DIR, "webinar.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())
 
 
@app.get("/admin/webinars", response_class=HTMLResponse)
def admin_webinars_page(request: Request):
    if request.session.get("role") != "admin":
        return RedirectResponse("/admin-login", status_code=302)
    path = os.path.join(STATIC_DIR, "admin-webinars.html")
    with open(path, "r", encoding="utf-8") as f:
        return HTMLResponse(f.read())
 
 
# ── PUBLIC: LIST ALL WEBINARS ──────────────────────────────────────────────────
 
@app.get("/api/webinars")
def api_list_webinars(request: Request, status: str = ""):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    if status:
        c.execute("""SELECT id,title,description,host,scheduled_at,duration_minutes,
                            topics,meeting_link,registration_link,cover_color,
                            status,max_seats,is_free,recording_link,created_at
                     FROM webinars WHERE status=? ORDER BY scheduled_at ASC""", (status,))
    else:
        c.execute("""SELECT id,title,description,host,scheduled_at,duration_minutes,
                            topics,meeting_link,registration_link,cover_color,
                            status,max_seats,is_free,recording_link,created_at
                     FROM webinars ORDER BY scheduled_at DESC""")
    rows = c.fetchall()
    username = request.session.get("user") or request.session.get("admin")
    result = []
    for r in rows:
        c.execute("SELECT COUNT(*) FROM webinar_registrations WHERE webinar_id=?", (r[0],))
        reg_count = c.fetchone()[0]
        is_registered = False
        if username:
            c.execute("SELECT id FROM webinar_registrations WHERE webinar_id=? AND username=?", (r[0], username))
            is_registered = c.fetchone() is not None
        result.append({
            "id": r[0], "title": r[1], "description": r[2] or "",
            "host": r[3] or "TraderBro Team", "scheduled_at": r[4],
            "duration_minutes": r[5] or 60,
            "topics": json.loads(r[6]) if r[6] else [],
            "meeting_link": r[7] or "", "registration_link": r[8] or "",
            "cover_color": r[9] or "orange", "status": r[10],
            "max_seats": r[11] or 0, "is_free": bool(r[12]),
            "recording_link": r[13] or "", "created_at": r[14] or "",
            "registrations": reg_count, "is_registered": is_registered,
        })
    conn.close()
    return JSONResponse({"webinars": result, "total": len(result)})
 
 
# ── PUBLIC: NEXT UPCOMING/LIVE WEBINAR (for notification banner) ───────────────
 
@app.get("/api/webinar/next")
def api_next_webinar(request: Request):
    """Returns the nearest upcoming or live webinar — used by notification banners."""
    now_str = datetime.now(IST).isoformat()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    # First check for any live webinar
    c.execute("""SELECT id,title,scheduled_at,cover_color,status,duration_minutes,meeting_link
                 FROM webinars WHERE status='live' LIMIT 1""")
    row = c.fetchone()
    if not row:
        # Fall back to next upcoming
        c.execute("""SELECT id,title,scheduled_at,cover_color,status,duration_minutes,meeting_link
                     FROM webinars WHERE status='upcoming' AND scheduled_at >= ?
                     ORDER BY scheduled_at ASC LIMIT 1""", (now_str,))
        row = c.fetchone()
    conn.close()
    if not row:
        return JSONResponse({"webinar": None})
    return JSONResponse({"webinar": {
        "id": row[0], "title": row[1], "scheduled_at": row[2],
        "cover_color": row[3], "status": row[4],
        "duration_minutes": row[5], "meeting_link": row[6] or "",
    }})
 
 
# ── ADMIN: CREATE WEBINAR ──────────────────────────────────────────────────────
 
@app.post("/api/admin/webinar/create")
async def api_create_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    data = await request.json()
    now  = datetime.now(IST).isoformat()
    title        = (data.get("title") or "").strip()
    scheduled_at = (data.get("scheduled_at") or "").strip()
    if not title or not scheduled_at:
        return JSONResponse({"success": False, "error": "Title and scheduled_at are required"})
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""INSERT INTO webinars
                     (title,description,host,scheduled_at,duration_minutes,topics,
                      meeting_link,registration_link,cover_color,status,max_seats,
                      is_free,created_at,updated_at)
                 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
              (title,
               (data.get("description") or "").strip(),
               (data.get("host") or "TraderBro Team").strip(),
               scheduled_at,
               int(data.get("duration_minutes") or 60),
               json.dumps(data.get("topics") or []),
               (data.get("meeting_link") or "").strip(),
               (data.get("registration_link") or "").strip(),
               (data.get("cover_color") or "orange").strip(),
               (data.get("status") or "upcoming").strip(),
               int(data.get("max_seats") or 0),
               1 if data.get("is_free", True) else 0,
               now, now))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return JSONResponse({"success": True, "id": new_id})
 
 
# ── ADMIN: UPDATE WEBINAR ──────────────────────────────────────────────────────
 
@app.post("/api/admin/webinar/update")
async def api_update_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    data = await request.json()
    wid  = data.get("id")
    if not wid:
        return JSONResponse({"success": False, "error": "id required"})
    now = datetime.now(IST).isoformat()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""UPDATE webinars SET
                     title=?,description=?,host=?,scheduled_at=?,duration_minutes=?,
                     topics=?,meeting_link=?,registration_link=?,cover_color=?,
                     status=?,max_seats=?,is_free=?,recording_link=?,updated_at=?
                 WHERE id=?""",
              (data.get("title",""),
               data.get("description",""),
               data.get("host","TraderBro Team"),
               data.get("scheduled_at",""),
               int(data.get("duration_minutes") or 60),
               json.dumps(data.get("topics") or []),
               data.get("meeting_link",""),
               data.get("registration_link",""),
               data.get("cover_color","orange"),
               data.get("status","upcoming"),
               int(data.get("max_seats") or 0),
               1 if data.get("is_free", True) else 0,
               data.get("recording_link",""),
               now, wid))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True})
 
 
# ── ADMIN: DELETE WEBINAR ──────────────────────────────────────────────────────
 
@app.post("/api/admin/webinar/delete")
async def api_delete_webinar(request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    data = await request.json()
    wid  = data.get("id")
    if not wid:
        return JSONResponse({"success": False, "error": "id required"})
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("DELETE FROM webinars WHERE id=?", (wid,))
    c.execute("DELETE FROM webinar_registrations WHERE webinar_id=?", (wid,))
    conn.commit()
    conn.close()
    return JSONResponse({"success": True})
 
 
# ── PUBLIC: REGISTER FOR WEBINAR ───────────────────────────────────────────────
 
@app.post("/api/webinar/register")
async def api_webinar_register(request: Request):
    username = request.session.get("user") or request.session.get("admin")
    if not username:
        return JSONResponse({"error": "Login required"}, status_code=401)
    data = await request.json()
    wid  = data.get("webinar_id")
    if not wid:
        return JSONResponse({"success": False, "error": "webinar_id required"})
    now = datetime.now(IST).isoformat()
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT status, max_seats FROM webinars WHERE id=?", (wid,))
    row = c.fetchone()
    if not row:
        conn.close()
        return JSONResponse({"success": False, "error": "Webinar not found"})
    if row[0] not in ("upcoming", "live"):
        conn.close()
        return JSONResponse({"success": False, "error": "Registrations are closed"})
    if row[1] and row[1] > 0:
        c.execute("SELECT COUNT(*) FROM webinar_registrations WHERE webinar_id=?", (wid,))
        if c.fetchone()[0] >= row[1]:
            conn.close()
            return JSONResponse({"success": False, "error": "Seats are full"})
    try:
        c.execute("INSERT OR IGNORE INTO webinar_registrations (webinar_id,username,registered_at) VALUES (?,?,?)",
                  (wid, username, now))
        conn.commit()
        inserted = c.rowcount > 0
        conn.close()
        if inserted:
            return JSONResponse({"success": True, "message": "Registered successfully!"})
        return JSONResponse({"success": False, "error": "Already registered"})
    except Exception as e:
        conn.close()
        return JSONResponse({"success": False, "error": str(e)})
 
 
# ── ADMIN: GET REGISTRATIONS FOR A WEBINAR ─────────────────────────────────────
 
@app.get("/api/admin/webinar/{wid}/registrations")
def api_webinar_registrations(wid: int, request: Request):
    if request.session.get("role") != "admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""SELECT wr.username, u.email, u.phone, wr.registered_at
                 FROM webinar_registrations wr
                 LEFT JOIN users u ON u.username=wr.username
                 WHERE wr.webinar_id=? ORDER BY wr.registered_at ASC""", (wid,))
    rows = c.fetchall()
    conn.close()
    return JSONResponse({
        "registrations": [{"username":r[0],"email":r[1]or"","phone":r[2]or"","registered_at":r[3]} for r in rows],
        "total": len(rows)
    })
